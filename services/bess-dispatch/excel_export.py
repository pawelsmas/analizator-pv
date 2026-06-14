"""
Excel Export Module for BESS Economics Analysis.

Generates detailed hourly breakdown Excel file with:
- For each hour: energia czynna (per ToU zone), dystrybucja, jakość, OZE, kog, akcyza, mocowa
- Separate columns for baseline (PV only) and project (PV+BESS)
- Full cost calculation per hour
- Summary sheet with totals

IMPORTANT - MVP Implementation:
1. Capacity fee (opłata mocowa) is calculated dynamically via capacity_fee_pl module
   using K-class classification (A=0.17/0.50/0.83/1.00) based on consumption profile.
   WOM in hourly rows is allocated from daily totals proportionally to import in selected hours.

2. OSD_ALL_IN mode: When tariff rates already include distribution (energia + dystrybucja),
   the distribution component should be set to 0 to avoid double counting.

3. All time calculations use CET_FIXED (UTC+1) to match Polish tariff definitions.

Version: 2.0.0
"""

import io
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from common.time_utils import CET_FIXED_OFFSET
from common.calendar_pl import get_day_type, DayType, is_polish_holiday
from capacity_fee_pl.models import CapacityFeeConfig, KClass, K_COEFFICIENTS, K_THRESHOLDS
from capacity_fee_pl.calculator import (
    compute_capacity_fee,
    compute_capacity_fee_savings,
    build_selected_hours_mask,
)


# =============================================================================
# Configuration Models
# =============================================================================

class FixedChargesConfig:
    """
    Fixed charges configuration (PLN/MWh).

    Note on OSD_ALL_IN mode:
    When is_osd_all_in=True, the ToU tariff rates already include distribution,
    so distribution should be set to 0 to avoid double counting.
    """
    def __init__(
        self,
        distribution: float = 200.0,
        distribution_peak: float = None,
        distribution_day: float = None,
        distribution_night: float = None,
        distribution_valley: float = None,
        quality_fee: float = 10.0,
        oze_fee: float = 7.0,
        cogeneration_fee: float = 10.0,
        excise_tax: float = 5.0,
        capacity_fee_som: float = 0.2194,  # SOM rate PLN/kWh
        is_osd_all_in: bool = False,  # If True, ToU rates include distribution
        # Distribution time windows (OSD zones — separate from energy ToU)
        dist_zone_type: str = 'three_zone',  # 'flat', 'two_zone', 'three_zone', 'four_zone'
        dist_two_zone_weekday_start: int = 6,
        dist_two_zone_weekday_end: int = 22,
        dist_two_zone_weekend_start: int = 6,
        dist_two_zone_weekend_end: int = 22,
        dist_peak1_start: int = 7,
        dist_peak1_end: int = 13,
        dist_peak2_start: int = 16,
        dist_peak2_end: int = 21,
        dist_weekend_off_peak: bool = True,
        dist_valley_start: int = 1,
        dist_valley_end: int = 5,
    ):
        self.quality_fee = quality_fee     # Opłata jakościowa
        self.oze_fee = oze_fee             # Opłata OZE
        self.cogeneration_fee = cogeneration_fee  # Opłata kogeneracyjna
        self.excise_tax = excise_tax       # Akcyza
        self.capacity_fee_som = capacity_fee_som  # SOM rate for capacity fee
        self.is_osd_all_in = is_osd_all_in

        # Distribution time windows (OSD zones)
        self.dist_zone_type = dist_zone_type
        self.dist_two_zone_weekday_start = dist_two_zone_weekday_start
        self.dist_two_zone_weekday_end = dist_two_zone_weekday_end
        self.dist_two_zone_weekend_start = dist_two_zone_weekend_start
        self.dist_two_zone_weekend_end = dist_two_zone_weekend_end
        self.dist_peak1_start = dist_peak1_start
        self.dist_peak1_end = dist_peak1_end
        self.dist_peak2_start = dist_peak2_start
        self.dist_peak2_end = dist_peak2_end
        self.dist_weekend_off_peak = dist_weekend_off_peak
        self.dist_valley_start = dist_valley_start
        self.dist_valley_end = dist_valley_end

        # If OSD_ALL_IN, distribution is already in ToU rates
        if is_osd_all_in:
            self.distribution = 0.0
            self.distribution_peak = 0.0
            self.distribution_day = 0.0
            self.distribution_night = 0.0
            self.distribution_valley = 0.0
        else:
            self.distribution = distribution  # Dystrybucja (weighted avg)
            # Zonal distribution: fallback to flat if not provided
            self.distribution_peak = distribution_peak if distribution_peak is not None else distribution
            self.distribution_day = distribution_day if distribution_day is not None else distribution
            self.distribution_night = distribution_night if distribution_night is not None else distribution
            self.distribution_valley = distribution_valley if distribution_valley is not None else (distribution_night if distribution_night is not None else distribution)

    @property
    def other_fees_total(self) -> float:
        """Sum of OTHER fees (excluding distribution and capacity fee)."""
        return (
            self.quality_fee +
            self.oze_fee +
            self.cogeneration_fee +
            self.excise_tax
        )

    @property
    def fixed_without_capacity(self) -> float:
        """Sum of fixed charges excluding capacity fee."""
        return self.distribution + self.other_fees_total


class ToUConfig:
    """Time-of-Use tariff configuration."""
    def __init__(
        self,
        tariff_type: str = "two_zone",  # flat, two_zone, three_zone
        flat_rate: float = 750.0,       # PLN/MWh
        day_rate: float = 850.0,        # PLN/MWh (two_zone)
        night_rate: float = 450.0,      # PLN/MWh (two_zone)
        peak_rate: float = 950.0,       # PLN/MWh (three_zone)
        partial_rate: float = 700.0,    # PLN/MWh (three_zone)
        off_peak_rate: float = 400.0,   # PLN/MWh (three_zone)
        weekday_day_start: int = 6,
        weekday_day_end: int = 22,
        weekend_day_start: int = 6,
        weekend_day_end: int = 13,
        peak1_start: int = 7,
        peak1_end: int = 13,
        peak2_start: int = 17,
        peak2_end: int = 21,
    ):
        self.tariff_type = tariff_type
        self.flat_rate = flat_rate
        self.day_rate = day_rate
        self.night_rate = night_rate
        self.peak_rate = peak_rate
        self.partial_rate = partial_rate
        self.off_peak_rate = off_peak_rate
        self.weekday_day_start = weekday_day_start
        self.weekday_day_end = weekday_day_end
        self.weekend_day_start = weekend_day_start
        self.weekend_day_end = weekend_day_end
        self.peak1_start = peak1_start
        self.peak1_end = peak1_end
        self.peak2_start = peak2_start
        self.peak2_end = peak2_end


# =============================================================================
# Time Index Builder (CET_FIXED)
# =============================================================================

def build_time_index_cet_fixed(
    start_date: date,
    n_timesteps: int,
    interval_minutes: int = 60
) -> List[datetime]:
    """
    Build a time index using CET_FIXED (UTC+1) for Polish tariff calculations.

    This ensures tariff zones are correctly applied according to Polish regulations
    which define zones in "czas zimowy" (winter time = CET).

    Args:
        start_date: Start date for the index
        n_timesteps: Number of timesteps
        interval_minutes: Interval duration (60 for hourly)

    Returns:
        List of datetime objects in CET_FIXED timezone
    """
    time_index = []
    dt = datetime(
        start_date.year, start_date.month, start_date.day,
        0, 0, 0, tzinfo=CET_FIXED_OFFSET
    )
    delta = timedelta(minutes=interval_minutes)

    for _ in range(n_timesteps):
        time_index.append(dt)
        dt += delta

    return time_index


# =============================================================================
# Zone Detection
# =============================================================================

def get_energia_zone(
    dt: datetime,
    tou_config: ToUConfig,
) -> Tuple[str, float]:
    """
    Determine energia czynna zone and rate for a given datetime.

    Args:
        dt: Datetime in CET-fixed timezone
        tou_config: ToU configuration

    Returns:
        Tuple of (zone_name, rate_pln_mwh)
    """
    hour = dt.hour
    weekday = dt.weekday()  # 0=Monday, 6=Sunday
    is_weekend = weekday >= 5

    if tou_config.tariff_type == "flat":
        return "Jednolita", tou_config.flat_rate

    elif tou_config.tariff_type == "two_zone":
        # Two-zone: day/night
        if is_weekend:
            # Weekend hours
            if tou_config.weekend_day_start <= hour < tou_config.weekend_day_end:
                return "Dzień", tou_config.day_rate
            else:
                return "Noc", tou_config.night_rate
        else:
            # Weekday hours
            if tou_config.weekday_day_start <= hour < tou_config.weekday_day_end:
                return "Dzień", tou_config.day_rate
            else:
                return "Noc", tou_config.night_rate

    elif tou_config.tariff_type == "three_zone":
        # Three-zone: peak/partial/off-peak (weekdays only for peak/partial)
        if is_weekend:
            return "Poza szczyt", tou_config.off_peak_rate
        else:
            # Check peak hours (morning and evening)
            if tou_config.peak1_start <= hour < tou_config.peak1_end:
                return "Szczyt", tou_config.peak_rate
            elif tou_config.peak2_start <= hour < tou_config.peak2_end:
                return "Szczyt", tou_config.peak_rate
            # Check partial hours (between peaks)
            elif tou_config.peak1_end <= hour < tou_config.peak2_start:
                return "Częściowy szczyt", tou_config.partial_rate
            else:
                return "Poza szczyt", tou_config.off_peak_rate

    # Default
    return "Nieznana", tou_config.flat_rate


def get_distribution_rate(
    dt: datetime,
    tou_config: ToUConfig,
    fixed_config: FixedChargesConfig,
) -> float:
    """
    Get distribution rate for a given hour based on OSD distribution time windows.

    Uses fixed_config.dist_* fields (separate from energy ToU zones).
    Returns distribution rate in PLN/MWh.
    """
    h = dt.hour
    is_weekend = dt.weekday() >= 5
    zone_type = fixed_config.dist_zone_type

    if zone_type == 'flat':
        return fixed_config.distribution_night  # flat = all same

    if zone_type == 'two_zone':
        if is_weekend:
            day_start = fixed_config.dist_two_zone_weekend_start
            day_end = fixed_config.dist_two_zone_weekend_end
        else:
            day_start = fixed_config.dist_two_zone_weekday_start
            day_end = fixed_config.dist_two_zone_weekday_end
        if day_start <= h < day_end:
            return fixed_config.distribution_day
        return fixed_config.distribution_night

    if zone_type == 'four_zone':
        # Weekend = full valley (Strefa 4)
        if is_weekend:
            return fixed_config.distribution_valley
        # Weekday: valley = deep night only (e.g. 1:00-4:59)
        v_start = fixed_config.dist_valley_start
        v_end = fixed_config.dist_valley_end
        if v_start <= h < v_end:
            return fixed_config.distribution_valley
        # Peak zones (Strefa 1+2)
        p1s = fixed_config.dist_peak1_start
        p1e = fixed_config.dist_peak1_end
        p2s = fixed_config.dist_peak2_start
        p2e = fixed_config.dist_peak2_end
        if (p1s <= h < p1e) or (p2s <= h < p2e):
            return fixed_config.distribution_peak
        # Remaining hours = Strefa 3 (fall through: 13-16, 21-1, 5-7)
        return fixed_config.distribution_day

    # three_zone (default)
    if is_weekend and fixed_config.dist_weekend_off_peak:
        return fixed_config.distribution_night
    p1s = fixed_config.dist_peak1_start
    p1e = fixed_config.dist_peak1_end
    p2s = fixed_config.dist_peak2_start
    p2e = fixed_config.dist_peak2_end
    if (p1s <= h < p1e) or (p2s <= h < p2e):
        return fixed_config.distribution_peak
    if h < p1s or h >= p2e:
        return fixed_config.distribution_night
    return fixed_config.distribution_day  # between peaks


def is_capacity_fee_hour(dt: datetime) -> bool:
    """
    Check if capacity fee (opłata mocowa) applies at given hour.

    Capacity fee applies:
    - Only on working days (Mon-Fri, excluding Polish holidays)
    - Hours 7:00-21:59 (7-22)

    Args:
        dt: Datetime to check (should have date info)

    Returns:
        True if capacity fee applies
    """
    # Check if weekend
    if dt.weekday() >= 5:  # Saturday=5, Sunday=6
        return False

    # Check if Polish holiday
    if is_polish_holiday(dt.date()):
        return False

    # Check hours (7:00-21:59)
    if 7 <= dt.hour < 22:
        return True

    return False


# =============================================================================
# Capacity Fee Calculator (Dynamic K-class)
# =============================================================================

def calculate_capacity_fee_dynamic(
    import_kw: np.ndarray,
    time_index: List[datetime],
    som_pln_kwh: float,
    interval_minutes: int = 60,
) -> Tuple[float, Dict[str, Any], np.ndarray]:
    """
    Calculate capacity fee using dynamic K-class classification.

    This uses the capacity_fee_pl module to properly calculate WOM with
    A coefficient based on consumption profile (Δs classification).

    Args:
        import_kw: Grid import profile [kW]
        time_index: List of datetime objects (CET_FIXED)
        som_pln_kwh: SOM rate [PLN/kWh]
        interval_minutes: Time resolution

    Returns:
        Tuple of (total_fee_pln, details_dict, hourly_allocation_pln)

    Note:
        hourly_allocation_pln contains the daily WOM allocated to each hour
        proportionally to import in selected hours. Sum of allocations = total WOM.
    """
    dt_hours = interval_minutes / 60.0
    n_timesteps = len(import_kw)

    # Convert kW to kWh
    import_kwh = import_kw * dt_hours

    # Build capacity fee config
    cap_config = CapacityFeeConfig(
        year=time_index[0].year if time_index else 2025,
        som_pln_per_kwh=som_pln_kwh,
    )

    # Calculate using capacity_fee_pl module
    result = compute_capacity_fee(
        grid_import_kwh=import_kwh,
        time_index=time_index,
        config=cap_config,
    )

    # Build selected hours mask for allocation
    selected_mask = build_selected_hours_mask(time_index, cap_config)

    # Allocate daily WOM to hourly values proportionally to import in selected hours
    hourly_allocation = np.zeros(n_timesteps)

    # Group by day
    daily_data: Dict[date, Dict] = {}
    for i, dt in enumerate(time_index):
        day = dt.date()
        if day not in daily_data:
            daily_data[day] = {'indices': [], 'selected_indices': []}
        daily_data[day]['indices'].append(i)
        if selected_mask[i]:
            daily_data[day]['selected_indices'].append(i)

    # Find daily fees from result
    daily_fees = {r.date: r.fee_pln for r in result.top_10_days}

    # For days not in top_10, we need to check monthly results
    # Actually, the result contains full daily breakdown through monthly_results
    # but individual day fees are in result.top_10_days only
    # Let's recalculate daily fees properly

    for day, day_info in daily_data.items():
        selected_indices = day_info['selected_indices']

        if not selected_indices:
            continue  # Non-workday or no selected hours

        # Get import in selected hours for this day
        selected_import_kwh = import_kwh[selected_indices]
        total_selected_kwh = float(np.sum(selected_import_kwh))

        if total_selected_kwh <= 0:
            continue

        # Calculate day's fee using K-class
        # We need to get the A coefficient for this specific day
        day_indices = day_info['indices']
        day_energy = import_kwh[day_indices]
        day_mask = selected_mask[day_indices]

        # Calculate Δs for this day
        zs = float(np.sum(day_energy[day_mask]))  # Energy in selected hours
        zps = float(np.sum(day_energy[~day_mask]))  # Energy outside selected hours
        n_selected = int(np.sum(day_mask))
        n_outside = int(np.sum(~day_mask))

        # Classify
        if zps == 0 or n_outside == 0:
            a_coeff = 1.0  # K4
        elif n_selected == 0:
            a_coeff = 1.0  # K4
        else:
            avg_s = zs / n_selected
            avg_ps = zps / n_outside
            if avg_ps == 0:
                a_coeff = 1.0
            else:
                delta_s = (avg_s / avg_ps - 1) * 100
                # Classify using K_THRESHOLDS from models.py
                a_coeff = 1.00  # K4 default
                for kc in [KClass.K1, KClass.K2, KClass.K3, KClass.K4]:
                    low, high = K_THRESHOLDS[kc]
                    if low <= delta_s < high:
                        a_coeff = K_COEFFICIENTS[kc]
                        break

        # Calculate day fee
        day_fee = a_coeff * som_pln_kwh * zs

        # Allocate proportionally to import in selected hours
        for idx in selected_indices:
            if total_selected_kwh > 0:
                proportion = import_kwh[idx] / total_selected_kwh
                hourly_allocation[idx] = day_fee * proportion

    details = {
        'total_fee_pln': result.total_fee_pln,
        'total_zs_kwh': result.total_zs_kwh,
        'k_histogram': result.k_histogram,
        'avg_delta_s': result.avg_delta_s,
        'workdays_count': result.workdays_count,
        'monthly_results': [
            {
                'month': m.month,
                'fee_pln': m.total_fee_pln,
                'zs_kwh': m.total_zs_kwh,
                'k_histogram': m.k_histogram,
            }
            for m in result.monthly_results
        ],
    }

    return result.total_fee_pln, details, hourly_allocation


# =============================================================================
# Hourly Cost Calculator
# =============================================================================

def calculate_hourly_costs(
    import_kw: np.ndarray,
    tou_config: ToUConfig,
    fixed_config: FixedChargesConfig,
    start_date: date,
    interval_minutes: int = 60,
    import_prices_pln_mwh: Optional[np.ndarray] = None,
    monthly_price_sources: Optional[Dict[int, str]] = None,
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Calculate detailed hourly costs with full breakdown.

    Uses capacity_fee_pl for dynamic K-class WOM calculation.

    Args:
        import_kw: Grid import power profile [kW]
        tou_config: Time-of-use configuration
        fixed_config: Fixed charges configuration
        start_date: Start date for analysis
        interval_minutes: Time resolution (60 for hourly)

    Returns:
        Tuple of (DataFrame with hourly breakdown, capacity_fee_details dict)
    """
    n_timesteps = len(import_kw)
    dt_hours = interval_minutes / 60

    # Build CET_FIXED time index
    time_index = build_time_index_cet_fixed(start_date, n_timesteps, interval_minutes)

    # Calculate capacity fee dynamically
    total_capacity_fee, cap_fee_details, hourly_cap_fee_allocation = calculate_capacity_fee_dynamic(
        import_kw=import_kw,
        time_index=time_index,
        som_pln_kwh=fixed_config.capacity_fee_som,
        interval_minutes=interval_minutes,
    )

    # Build hourly records
    records = []

    for i, dt in enumerate(time_index):
        import_kwh = import_kw[i] * dt_hours
        import_mwh = import_kwh / 1000.0

        # Get energia czynna zone and rate
        zone_name, tou_rate = get_energia_zone(dt, tou_config)

        # Determine price source for this hour
        month = dt.month
        if monthly_price_sources:
            # Hybrid monthly: check per-month source
            source = monthly_price_sources.get(month, monthly_price_sources.get(str(month), 'osd'))
            if source == 'rdn' and import_prices_pln_mwh is not None and i < len(import_prices_pln_mwh):
                energia_rate = float(import_prices_pln_mwh[i])
                zone_name = f"RDN ({energia_rate:.0f})"
            else:
                energia_rate = tou_rate
        elif import_prices_pln_mwh is not None and i < len(import_prices_pln_mwh):
            # Full-year RDN
            energia_rate = float(import_prices_pln_mwh[i])
            zone_name = f"RDN ({energia_rate:.0f})"
        else:
            energia_rate = tou_rate
        energia_pln = import_mwh * energia_rate

        # Fixed charges (distribution zonal — may be 0 if OSD_ALL_IN)
        dist_rate = get_distribution_rate(dt, tou_config, fixed_config)
        dystrybucja_pln = import_mwh * dist_rate
        jakosc_pln = import_mwh * fixed_config.quality_fee
        oze_pln = import_mwh * fixed_config.oze_fee
        kog_pln = import_mwh * fixed_config.cogeneration_fee
        akcyza_pln = import_mwh * fixed_config.excise_tax

        # Capacity fee from allocation (already calculated with K-class)
        mocowa_pln = hourly_cap_fee_allocation[i]

        # Calculate effective hourly rate for display (if applicable)
        is_selected_hour = is_capacity_fee_hour(dt)
        if is_selected_hour and import_kwh > 0:
            mocowa_rate_effective = mocowa_pln / import_mwh if import_mwh > 0 else 0
        else:
            mocowa_rate_effective = 0.0

        # Total
        total_pln = (
            energia_pln + dystrybucja_pln + jakosc_pln +
            oze_pln + kog_pln + akcyza_pln + mocowa_pln
        )

        # Effective rate (total cost / import)
        if import_mwh > 0:
            effective_rate = total_pln / import_mwh
        else:
            effective_rate = energia_rate + fixed_config.fixed_without_capacity

        records.append({
            'datetime': dt.replace(tzinfo=None),  # Excel doesn't support timezone
            'date': dt.date(),
            'hour': dt.hour,
            'weekday': dt.strftime('%A'),
            'weekday_pl': ['Pon', 'Wt', 'Śr', 'Czw', 'Pt', 'Sob', 'Nd'][dt.weekday()],
            'is_holiday': is_polish_holiday(dt.date()),
            'is_capacity_hour': is_selected_hour,
            'zone': zone_name,
            'import_kwh': round(import_kwh, 3),
            'import_mwh': round(import_mwh, 6),
            # Rates (PLN/MWh)
            'energia_rate': energia_rate,
            'dystrybucja_rate': dist_rate,
            'jakosc_rate': fixed_config.quality_fee,
            'oze_rate': fixed_config.oze_fee,
            'kog_rate': fixed_config.cogeneration_fee,
            'akcyza_rate': fixed_config.excise_tax,
            'mocowa_rate': round(mocowa_rate_effective, 2),  # Effective rate for this hour
            'effective_rate': round(effective_rate, 2),
            # Costs (PLN)
            'energia_pln': round(energia_pln, 4),
            'dystrybucja_pln': round(dystrybucja_pln, 4),
            'jakosc_pln': round(jakosc_pln, 4),
            'oze_pln': round(oze_pln, 4),
            'kog_pln': round(kog_pln, 4),
            'akcyza_pln': round(akcyza_pln, 4),
            'mocowa_pln': round(mocowa_pln, 4),  # This is ALLOCATED from daily WOM
            'total_pln': round(total_pln, 4),
        })

    return pd.DataFrame(records), cap_fee_details


# =============================================================================
# Excel Generator
# =============================================================================

def generate_economics_excel(
    baseline_import_kw: np.ndarray,
    project_import_kw: np.ndarray,
    tou_config: ToUConfig,
    fixed_config: FixedChargesConfig,
    start_date: date,
    bess_power_kw: float,
    bess_energy_kwh: float,
    interval_minutes: int = 60,
    project_name: str = "Analiza BESS",
    **kwargs,
) -> bytes:
    """
    Generate Excel with LIVE FORMULAS for full transparency.

    Static data (from solver): import profiles, battery trace, capacity fee allocation, RDN prices.
    Formulas: all cost calculations, totals, savings, summaries.
    """
    import_prices_pln_mwh = kwargs.get('import_prices_pln_mwh', None)
    if import_prices_pln_mwh is not None:
        import_prices_pln_mwh = np.asarray(import_prices_pln_mwh, dtype=float)

    monthly_price_sources = kwargs.get('monthly_price_sources', None)
    # Normalize keys to int if provided
    if monthly_price_sources:
        monthly_price_sources = {int(k): v for k, v in monthly_price_sources.items()}

    # We still need capacity fee from Python (K-class is too complex for Excel formulas)
    baseline_df, baseline_cap_details = calculate_hourly_costs(
        baseline_import_kw, tou_config, fixed_config, start_date, interval_minutes,
        import_prices_pln_mwh=import_prices_pln_mwh,
        monthly_price_sources=monthly_price_sources,
    )
    project_df, project_cap_details = calculate_hourly_costs(
        project_import_kw, tou_config, fixed_config, start_date, interval_minutes,
        import_prices_pln_mwh=import_prices_pln_mwh,
        monthly_price_sources=monthly_price_sources,
    )

    n_timesteps = len(baseline_import_kw)
    dt_hours = interval_minutes / 60
    period_days = int(np.ceil(n_timesteps * dt_hours / 24))
    end_date = start_date + timedelta(days=period_days - 1)
    time_index = build_time_index_cet_fixed(start_date, n_timesteps, interval_minutes)

    # Pre-extract capacity fee hourly allocations (static - too complex for formula)
    bl_mocowa = baseline_df['mocowa_pln'].values if 'mocowa_pln' in baseline_df.columns else np.zeros(n_timesteps)
    pj_mocowa = project_df['mocowa_pln'].values if 'mocowa_pln' in project_df.columns else np.zeros(n_timesteps)

    has_rdn = import_prices_pln_mwh is not None and len(import_prices_pln_mwh) >= n_timesteps
    is_pv_only = bess_power_kw < 0.1 and bess_energy_kwh < 0.1

    # For hybrid monthly: determine which hours use RDN vs OSD
    def _get_hour_price_and_zone(dt_cet, idx):
        """Return (zone_name, price_pln_mwh) respecting hybrid monthly sources."""
        month = dt_cet.month
        if monthly_price_sources:
            source = monthly_price_sources.get(month, monthly_price_sources.get(str(month), 'osd'))
            if source == 'rdn' and has_rdn:
                price = float(import_prices_pln_mwh[idx])
                return f"RDN ({price:.0f})", price
            else:
                zone_name, rate = get_energia_zone(dt_cet, tou_config)
                return zone_name, rate
        elif has_rdn:
            price = float(import_prices_pln_mwh[idx])
            return f"RDN ({price:.0f})", price
        else:
            return get_energia_zone(dt_cet, tou_config)

    wb = Workbook()

    # ====================== STYLES ======================
    hdr_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    sec_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    sec_font = Font(bold=True, color="1565C0")
    warn_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    nfmt = '#,##0.00'
    nfmt4 = '#,##0.0000'
    batt_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    charge_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    discharge_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    pv_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    formula_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    savings_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    # =====================================================================
    # Sheet 1: Podsumowanie (with formulas referencing hourly sheet)
    # =====================================================================
    ws = wb.active
    ws.title = "Podsumowanie"

    ws['A1'] = project_name
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    ws['A2'] = f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = f"Okres analizy: {start_date} - {end_date} ({period_days} dni)"

    if is_pv_only:
        pv_kw_arr = kwargs.get('pv_kw', None)
        pv_capacity = kwargs.get('pv_capacity_kwp', 0)
        if pv_capacity == 0 and pv_kw_arr is not None:
            pv_capacity = round(float(np.max(pv_kw_arr)), 1)
        ws['A5'] = "Konfiguracja PV"
        ws['A5'].font = sec_font; ws['A5'].fill = sec_fill
        ws.merge_cells('A5:D5')
        ws['A6'] = "Moc PV [kWp]"; ws['B6'] = pv_capacity
    else:
        ws['A5'] = "Konfiguracja magazynu"
        ws['A5'].font = sec_font; ws['A5'].fill = sec_fill
        ws.merge_cells('A5:D5')
        ws['A6'] = "Moc [kW]"; ws['B6'] = bess_power_kw
        ws['A7'] = "Pojemność [kWh]"; ws['B7'] = bess_energy_kwh

    ws['A9'] = "Stawki energii"
    ws['A9'].font = sec_font; ws['A9'].fill = sec_fill
    ws.merge_cells('A9:D9')

    if monthly_price_sources:
        ws['A10'] = "Źródło cen"
        osd_months = [m for m, s in monthly_price_sources.items() if s == 'osd']
        rdn_months = [m for m, s in monthly_price_sources.items() if s == 'rdn']
        ws['B10'] = f"Miks: OSD ({len(osd_months)} mies.) + RDN ({len(rdn_months)} mies.)"
        ws['B10'].font = Font(bold=True, color="1565C0")
        ws['A11'] = "Miesiące OSD"; ws['B11'] = ', '.join(str(m) for m in sorted(osd_months))
        ws['A12'] = "Miesiące RDN"; ws['B12'] = ', '.join(str(m) for m in sorted(rdn_months))
        ws['B12'].font = Font(bold=True, color="D32F2F")
    elif has_rdn:
        ws['A10'] = "Źródło cen"
        ws['B10'] = "RDN (ceny godzinowe z rynku)"
        ws['B10'].font = Font(bold=True, color="D32F2F")
    else:
        ws['A10'] = "Typ taryfy"
        ws['B10'] = tou_config.tariff_type
        if fixed_config.is_osd_all_in:
            ws['C10'] = "Stawki zawierają dystrybucję"
            ws['C10'].fill = warn_fill

    r = 11
    if not has_rdn:
        if tou_config.tariff_type == "two_zone":
            ws[f'A{r}'] = "Stawka dzienna [PLN/MWh]"; ws[f'B{r}'] = tou_config.day_rate; r += 1
            ws[f'A{r}'] = "Stawka nocna [PLN/MWh]"; ws[f'B{r}'] = tou_config.night_rate; r += 1
        elif tou_config.tariff_type == "three_zone":
            ws[f'A{r}'] = "Stawka szczyt [PLN/MWh]"; ws[f'B{r}'] = tou_config.peak_rate; r += 1
            ws[f'A{r}'] = "Stawka częściowy [PLN/MWh]"; ws[f'B{r}'] = tou_config.partial_rate; r += 1
            ws[f'A{r}'] = "Stawka poza szczyt [PLN/MWh]"; ws[f'B{r}'] = tou_config.off_peak_rate; r += 1

    r += 1
    ws[f'A{r}'] = "Opłaty dodatkowe [PLN/MWh]"
    ws[f'A{r}'].font = sec_font; ws[f'A{r}'].fill = sec_fill
    ws.merge_cells(f'A{r}:D{r}'); r += 1

    # Store rate rows for reference
    rate_rows = {}
    for name, val in [
        ("Dystrybucja szczyt", fixed_config.distribution_peak),
        ("Dystrybucja dzień", fixed_config.distribution_day),
        ("Dystrybucja noc", fixed_config.distribution_night),
        ("Opłata jakościowa", fixed_config.quality_fee),
        ("Opłata OZE", fixed_config.oze_fee),
        ("Opłata kogeneracyjna", fixed_config.cogeneration_fee),
        ("Akcyza", fixed_config.excise_tax),
    ]:
        ws[f'A{r}'] = name; ws[f'B{r}'] = val
        rate_rows[name] = r; r += 1

    ws[f'A{r}'] = "Opłata mocowa SOM [PLN/kWh]"
    ws[f'B{r}'] = fixed_config.capacity_fee_som
    ws[f'C{r}'] = "(klasyfikacja K - wyliczana dynamicznie)"
    r += 1

    # K-class summary
    r += 1
    ws[f'A{r}'] = "Opłata mocowa - klasyfikacja K"
    ws[f'A{r}'].font = sec_font; ws[f'A{r}'].fill = sec_fill
    ws.merge_cells(f'A{r}:D{r}'); r += 1

    bl_label = "Bez PV" if is_pv_only else "Bez magazynu"
    pj_label = "Z PV" if is_pv_only else "Z magazynem"
    for c, label in [(2, bl_label), (3, pj_label), (4, "Oszczędności")]:
        cell = ws.cell(row=r, column=c, value=label)
        cell.fill = hdr_fill; cell.font = hdr_font
    r += 1

    ws[f'A{r}'] = "WOM [PLN]"
    ws[f'B{r}'] = round(baseline_cap_details['total_fee_pln'], 2)
    ws[f'C{r}'] = round(project_cap_details['total_fee_pln'], 2)
    ws[f'D{r}'] = f'=B{r}-C{r}'; ws[f'D{r}'].number_format = nfmt; r += 1

    ws[f'A{r}'] = "Średnia Δs [%]"
    ws[f'B{r}'] = round(baseline_cap_details['avg_delta_s'], 1)
    ws[f'C{r}'] = round(project_cap_details['avg_delta_s'], 1); r += 1

    for klass, coeff in [("K1", 0.17), ("K2", 0.50), ("K3", 0.83), ("K4", 1.00)]:
        ws[f'A{r}'] = f"Dni {klass} (A={coeff})"
        ws[f'B{r}'] = baseline_cap_details['k_histogram'].get(klass, 0)
        ws[f'C{r}'] = project_cap_details['k_histogram'].get(klass, 0); r += 1

    # Annual summary with FORMULAS referencing hourly sheet
    r += 1
    ws[f'A{r}'] = "Podsumowanie roczne"
    ws[f'A{r}'].font = sec_font; ws[f'A{r}'].fill = sec_fill
    ws.merge_cells(f'A{r}:D{r}'); r += 1

    bl_sum_label = "Bez PV (pełny pobór)" if is_pv_only else "Bez magazynu (tylko PV)"
    pj_sum_label = "Z PV (z autokonsumpcją)" if is_pv_only else "Z magazynem (PV+BESS)"
    for c, label in [(1, ""), (2, bl_sum_label), (3, pj_sum_label), (4, "Oszczędności")]:
        cell = ws.cell(row=r, column=c, value=label)
        cell.fill = hdr_fill; cell.font = hdr_font
    r += 1
    sum_start = r

    # Column letters in hourly sheet: E=bl_import, N=bl_total, O=pj_import, W=pj_total
    # We'll define these after building the hourly sheet - for now use placeholder refs
    hs = "'Rozliczenie Godzinowe'"
    last_data_row = n_timesteps + 3  # header at row 3, data starts row 4

    summary_formulas = [
        ("Pobór energii [MWh]",
         f"=SUM({hs}!E4:E{last_data_row})/1000",
         f"=SUM({hs}!O4:O{last_data_row})/1000",
         f"=B{r}-C{r}"),
        ("Koszt energii [PLN]",
         f"=SUM({hs}!G4:G{last_data_row})",
         f"=SUM({hs}!P4:P{last_data_row})",
         f"=B{r+1}-C{r+1}"),
        ("Dystrybucja [PLN]",
         f"=SUM({hs}!H4:H{last_data_row})",
         f"=SUM({hs}!Q4:Q{last_data_row})",
         f"=B{r+2}-C{r+2}"),
        ("Opłata jakościowa [PLN]",
         f"=SUM({hs}!I4:I{last_data_row})",
         f"=SUM({hs}!R4:R{last_data_row})",
         f"=B{r+3}-C{r+3}"),
        ("Opłata OZE [PLN]",
         f"=SUM({hs}!J4:J{last_data_row})",
         f"=SUM({hs}!S4:S{last_data_row})",
         f"=B{r+4}-C{r+4}"),
        ("Opłata kogeneracyjna [PLN]",
         f"=SUM({hs}!K4:K{last_data_row})",
         f"=SUM({hs}!T4:T{last_data_row})",
         f"=B{r+5}-C{r+5}"),
        ("Akcyza [PLN]",
         f"=SUM({hs}!L4:L{last_data_row})",
         f"=SUM({hs}!U4:U{last_data_row})",
         f"=B{r+6}-C{r+6}"),
        ("Opłata mocowa [PLN]",
         f"=SUM({hs}!M4:M{last_data_row})",
         f"=SUM({hs}!V4:V{last_data_row})",
         f"=B{r+7}-C{r+7}"),
        ("RAZEM [PLN]",
         f"=SUM({hs}!N4:N{last_data_row})",
         f"=SUM({hs}!W4:W{last_data_row})",
         f"=B{r+8}-C{r+8}"),
    ]

    for i, (label, f_bl, f_pj, f_sav) in enumerate(summary_formulas):
        cr = r + i
        ws.cell(row=cr, column=1, value=label)
        ws.cell(row=cr, column=2, value=f_bl).number_format = nfmt
        ws.cell(row=cr, column=3, value=f_pj).number_format = nfmt
        ws.cell(row=cr, column=4, value=f_sav).number_format = nfmt
        if i == len(summary_formulas) - 1:
            for c in range(1, 5):
                ws.cell(row=cr, column=c).font = Font(bold=True)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20

    # =====================================================================
    # Sheet 2: Rozliczenie Godzinowe (WITH FORMULAS)
    # =====================================================================
    ws_h = wb.create_sheet("Rozliczenie Godzinowe")

    # Row 1: explanation
    if is_pv_only:
        ws_h['A1'] = ('Objaśnienie: Kolumny z fioletowym tłem zawierają FORMUŁY - możesz je zweryfikować. '
                      '"Bez PV" = pełny pobór z sieci, "Z PV" = pobór z uwzględnieniem autokonsumpcji PV.')
    else:
        ws_h['A1'] = ('Objaśnienie: Kolumny z fioletowym tłem zawierają FORMUŁY - możesz je zweryfikować. '
                      '"Bez mag." = tylko PV, "Z mag." = PV + magazyn energii.')
    ws_h['A1'].font = Font(italic=True, color="666666")
    ws_h.merge_cells('A1:X1')

    # Row 2: rate reference
    rate_info = f'Stawki [PLN/MWh]: Dystr.szczyt={fixed_config.distribution_peak}, Dystr.dzień={fixed_config.distribution_day}, Dystr.noc={fixed_config.distribution_night}, Jakość={fixed_config.quality_fee}, OZE={fixed_config.oze_fee}, Kogeneracja={fixed_config.cogeneration_fee}, Akcyza={fixed_config.excise_tax}'
    if monthly_price_sources:
        osd_m = sorted([m for m, s in monthly_price_sources.items() if s == 'osd'])
        rdn_m = sorted([m for m, s in monthly_price_sources.items() if s == 'rdn'])
        rate_info += f' | Miks cenowy: OSD mies.={osd_m}, RDN mies.={rdn_m}'
    ws_h['A2'] = rate_info
    ws_h['A2'].font = Font(italic=True, color="999999", size=9)
    ws_h.merge_cells('A2:X2')

    # Determine labels
    bl_tag = "Bez PV" if is_pv_only else "Bez mag."
    pj_tag = "Z PV" if is_pv_only else "Z mag."
    zone_hdr = "Strefa / Źródło ceny" if monthly_price_sources else ("Strefa / Cena RDN" if has_rdn else "Strefa taryfowa")

    # Headers row 3
    # A=datetime, B=weekday, C=capacity_hour, D=zone
    # E=bl_import_kwh, F=bl_price, G=bl_energia, H=bl_dystr, I=bl_jakosc, J=bl_oze, K=bl_kog, L=bl_akcyza, M=bl_mocowa, N=bl_total
    # O=pj_import_kwh, P=pj_energia, Q=pj_dystr, R=pj_jakosc, S=pj_oze, T=pj_kog, U=pj_akcyza, V=pj_mocowa, W=pj_total
    # X=savings

    headers = [
        ('Data i godzina', 18),
        ('Dzień tyg.', 10),
        ('Godz. mocowa', 10),
        (zone_hdr, 16),
        # Baseline
        (f'{bl_tag}\nPobór [kWh]', 14),
        (f'{bl_tag}\nCena energii\n[PLN/MWh]', 14),
        (f'{bl_tag}\nKoszt energii\n[PLN]', 14),
        (f'{bl_tag}\nDystrybucja\n[PLN]', 14),
        (f'{bl_tag}\nOpł. jakość\n[PLN]', 12),
        (f'{bl_tag}\nOpł. OZE\n[PLN]', 12),
        (f'{bl_tag}\nOpł. kogen.\n[PLN]', 12),
        (f'{bl_tag}\nAkcyza\n[PLN]', 12),
        (f'{bl_tag}\nOpł. mocowa\n[PLN]', 14),
        (f'{bl_tag}\nKOSZT RAZEM\n[PLN]', 14),
        # Project
        (f'{pj_tag}\nPobór [kWh]', 14),
        (f'{pj_tag}\nKoszt energii\n[PLN]', 14),
        (f'{pj_tag}\nDystrybucja\n[PLN]', 14),
        (f'{pj_tag}\nOpł. jakość\n[PLN]', 12),
        (f'{pj_tag}\nOpł. OZE\n[PLN]', 12),
        (f'{pj_tag}\nOpł. kogen.\n[PLN]', 12),
        (f'{pj_tag}\nAkcyza\n[PLN]', 12),
        (f'{pj_tag}\nOpł. mocowa\n[PLN]', 14),
        (f'{pj_tag}\nKOSZT RAZEM\n[PLN]', 14),
        # Savings
        ('OSZCZĘDNOŚĆ\n[PLN]', 14),
    ]

    # PV-only: add PV profile columns
    pv_kw_arr = kwargs.get('pv_kw', None)
    load_kw_arr = kwargs.get('load_kw', None)
    if is_pv_only and pv_kw_arr is not None:
        headers.extend([
            ('Konsumpcja\n[kWh]', 14),
            ('Produkcja PV\n[kWh]', 14),
            ('Autokonsumpcja\n[kWh]', 14),
            ('Nadwyżka PV\n[kWh]', 14),
        ])

    for ci, (h, w) in enumerate(headers, 1):
        cell = ws_h.cell(row=3, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws_h.column_dimensions[get_column_letter(ci)].width = w

    # Rate constants (PLN/MWh) for formulas
    # NOTE: dist_rate is per-hour (zonal), computed inside the loop
    qual_rate = fixed_config.quality_fee
    oze_rate = fixed_config.oze_fee
    kog_rate = fixed_config.cogeneration_fee
    exc_rate = fixed_config.excise_tax

    # Write data rows with FORMULAS
    for i in range(n_timesteps):
        r = i + 4  # data starts at row 4
        dt = time_index[i]

        # A: datetime
        ws_h.cell(row=r, column=1, value=dt.replace(tzinfo=None))
        # B: weekday
        ws_h.cell(row=r, column=2, value=['Pon', 'Wt', 'Śr', 'Czw', 'Pt', 'Sob', 'Nd'][dt.weekday()])
        # C: capacity hour
        ws_h.cell(row=r, column=3, value="TAK" if is_capacity_fee_hour(dt) else "")
        # D: zone/price (respects hybrid monthly)
        zone_name, price = _get_hour_price_and_zone(dt, i)
        ws_h.cell(row=r, column=4, value=zone_name)

        # E: Baseline - Pobór [kWh] (STATIC from solver)
        bl_kwh = float(baseline_import_kw[i]) * dt_hours
        ws_h.cell(row=r, column=5, value=round(bl_kwh, 4))

        # F: Baseline - Cena energii [PLN/MWh] (STATIC, respects hybrid monthly)
        ws_h.cell(row=r, column=6, value=round(price, 2))

        # Distribution rate for this hour (zonal)
        dist_rate_h = get_distribution_rate(dt, tou_config, fixed_config)

        # G: Bez mag - Koszt energii = E * F / 1000 (FORMULA!)
        c = ws_h.cell(row=r, column=7, value=f'=E{r}*F{r}/1000')
        c.number_format = nfmt4; c.fill = formula_fill

        # H: Bez mag - Dystrybucja = E / 1000 * rate (FORMULA with zonal rate!)
        c = ws_h.cell(row=r, column=8, value=f'=E{r}/1000*{dist_rate_h}')
        c.number_format = nfmt4; c.fill = formula_fill

        # I: Bez mag - Jakość (FORMULA)
        c = ws_h.cell(row=r, column=9, value=f'=E{r}/1000*{qual_rate}')
        c.number_format = nfmt4; c.fill = formula_fill

        # J: Bez mag - OZE (FORMULA)
        c = ws_h.cell(row=r, column=10, value=f'=E{r}/1000*{oze_rate}')
        c.number_format = nfmt4; c.fill = formula_fill

        # K: Bez mag - Kogeneracja (FORMULA)
        c = ws_h.cell(row=r, column=11, value=f'=E{r}/1000*{kog_rate}')
        c.number_format = nfmt4; c.fill = formula_fill

        # L: Bez mag - Akcyza (FORMULA)
        c = ws_h.cell(row=r, column=12, value=f'=E{r}/1000*{exc_rate}')
        c.number_format = nfmt4; c.fill = formula_fill

        # M: Bez mag - Opł. mocowa (STATIC - K-class too complex)
        ws_h.cell(row=r, column=13, value=round(float(bl_mocowa[i]), 4)).number_format = nfmt4

        # N: Bez mag - KOSZT RAZEM = SUM(G:M) (FORMULA!)
        c = ws_h.cell(row=r, column=14, value=f'=SUM(G{r}:M{r})')
        c.number_format = nfmt4; c.fill = formula_fill; c.font = Font(bold=True)

        # O: Z mag - Pobór [kWh] (STATIC from solver)
        pj_kwh = float(project_import_kw[i]) * dt_hours
        ws_h.cell(row=r, column=15, value=round(pj_kwh, 4))

        # P: Z mag - Koszt energii = O * F / 1000 (FORMULA - same price!)
        c = ws_h.cell(row=r, column=16, value=f'=O{r}*F{r}/1000')
        c.number_format = nfmt4; c.fill = formula_fill

        # Q: Z mag - Dystrybucja (FORMULA with zonal rate!)
        c = ws_h.cell(row=r, column=17, value=f'=O{r}/1000*{dist_rate_h}')
        c.number_format = nfmt4; c.fill = formula_fill

        # R: Z mag - Jakość (FORMULA)
        c = ws_h.cell(row=r, column=18, value=f'=O{r}/1000*{qual_rate}')
        c.number_format = nfmt4; c.fill = formula_fill

        # S: Z mag - OZE (FORMULA)
        c = ws_h.cell(row=r, column=19, value=f'=O{r}/1000*{oze_rate}')
        c.number_format = nfmt4; c.fill = formula_fill

        # T: Z mag - Kogeneracja (FORMULA)
        c = ws_h.cell(row=r, column=20, value=f'=O{r}/1000*{kog_rate}')
        c.number_format = nfmt4; c.fill = formula_fill

        # U: Z mag - Akcyza (FORMULA)
        c = ws_h.cell(row=r, column=21, value=f'=O{r}/1000*{exc_rate}')
        c.number_format = nfmt4; c.fill = formula_fill

        # V: Z mag - Opł. mocowa (STATIC)
        ws_h.cell(row=r, column=22, value=round(float(pj_mocowa[i]), 4)).number_format = nfmt4

        # W: Z mag - KOSZT RAZEM = SUM(P:V) (FORMULA!)
        c = ws_h.cell(row=r, column=23, value=f'=SUM(P{r}:V{r})')
        c.number_format = nfmt4; c.fill = formula_fill; c.font = Font(bold=True)

        # X: OSZCZĘDNOŚĆ = N - W (FORMULA!)
        c = ws_h.cell(row=r, column=24, value=f'=N{r}-W{r}')
        c.number_format = nfmt4; c.fill = savings_fill; c.font = Font(bold=True)

        # PV-only extra columns: Y=Load, Z=PV, AA=SelfConsumed, AB=Surplus
        if is_pv_only and pv_kw_arr is not None:
            load_val = float(load_kw_arr[i]) * dt_hours if load_kw_arr is not None and i < len(load_kw_arr) else 0
            pv_val = float(pv_kw_arr[i]) * dt_hours if i < len(pv_kw_arr) else 0
            ws_h.cell(row=r, column=25, value=round(load_val, 2)).number_format = nfmt
            ws_h.cell(row=r, column=26, value=round(pv_val, 2)).number_format = nfmt
            # Autokonsumpcja = MIN(PV, Load) (FORMULA)
            c = ws_h.cell(row=r, column=27, value=f'=MIN(Y{r},Z{r})')
            c.number_format = nfmt; c.fill = formula_fill
            # Nadwyżka PV = MAX(0, PV - Load) (FORMULA)
            c = ws_h.cell(row=r, column=28, value=f'=MAX(0,Z{r}-Y{r})')
            c.number_format = nfmt; c.fill = formula_fill

    ws_h.freeze_panes = 'A4'

    # =====================================================================
    # Sheet 3: Podsumowanie Miesięczne (FORMULAS with SUMPRODUCT)
    # =====================================================================
    ws_m = wb.create_sheet("Podsumowanie Miesięczne")

    m_headers = [
        'Miesiąc',
        f'{bl_tag}\nPobór [kWh]', f'{bl_tag}\nKoszt energii', f'{bl_tag}\nDystrybucja',
        f'{bl_tag}\nOpł. mocowa', f'{bl_tag}\nRazem [PLN]',
        f'{pj_tag}\nPobór [kWh]', f'{pj_tag}\nKoszt energii', f'{pj_tag}\nDystrybucja',
        f'{pj_tag}\nOpł. mocowa', f'{pj_tag}\nRazem [PLN]',
        'Oszczędność\nna energii', 'Oszczędność\nna mocowej',
        'Oszczędność\nRAZEM [PLN]',
    ]

    for ci, h in enumerate(m_headers, 1):
        cell = ws_m.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws_m.column_dimensions[get_column_letter(ci)].width = 18

    # Build monthly data using SUMPRODUCT formulas
    months_in_data = sorted(set(dt.month for dt in time_index))
    year = start_date.year

    for mi, month in enumerate(months_in_data):
        r = mi + 2
        month_names = ['Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec',
                       'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']
        ws_m.cell(row=r, column=1, value=f"{month_names[month-1]} {year}")

        # SUMPRODUCT to sum only rows where MONTH(A)=month
        # A column in hourly sheet is datetime
        hs_ref = "'Rozliczenie Godzinowe'"
        rng = f"$A$4:$A${last_data_row}"
        month_crit = f'(MONTH({hs_ref}!{rng})={month})'

        col_map = {
            2: 'E', 3: 'G', 4: 'H', 5: 'M', 6: 'N',    # Bez mag
            7: 'O', 8: 'P', 9: 'Q', 10: 'V', 11: 'W',   # Z mag
        }

        for col_idx, src_col in col_map.items():
            src_rng = f"${src_col}$4:${src_col}${last_data_row}"
            formula = f"=SUMPRODUCT({month_crit}*{hs_ref}!{src_rng})"
            c = ws_m.cell(row=r, column=col_idx, value=formula)
            c.number_format = nfmt; c.fill = formula_fill

        # Savings = differences (FORMULA)
        ws_m.cell(row=r, column=12, value=f'=C{r}-H{r}').number_format = nfmt  # energia savings
        ws_m.cell(row=r, column=12).fill = savings_fill
        ws_m.cell(row=r, column=13, value=f'=E{r}-J{r}').number_format = nfmt  # mocowa savings
        ws_m.cell(row=r, column=13).fill = savings_fill
        ws_m.cell(row=r, column=14, value=f'=F{r}-K{r}').number_format = nfmt  # total savings
        ws_m.cell(row=r, column=14).fill = savings_fill
        ws_m.cell(row=r, column=14).font = Font(bold=True)

    # Total row
    tr = len(months_in_data) + 2
    ws_m.cell(row=tr, column=1, value="RAZEM ROK").font = Font(bold=True)
    for ci in range(2, 15):
        col_letter = get_column_letter(ci)
        ws_m.cell(row=tr, column=ci, value=f'=SUM({col_letter}2:{col_letter}{tr-1})')
        ws_m.cell(row=tr, column=ci).number_format = nfmt
        ws_m.cell(row=tr, column=ci).font = Font(bold=True)
        ws_m.cell(row=tr, column=ci).fill = savings_fill

    ws_m.freeze_panes = 'A2'

    # =====================================================================
    # Sheet 4: Praca Magazynu (WITH FORMULAS)
    # =====================================================================
    battery_trace = kwargs.get('battery_trace', None)
    if battery_trace and isinstance(battery_trace, dict):
        ws_b = wb.create_sheet("Praca Magazynu")

        soc_arr = battery_trace.get('soc_kwh')
        charge_arr = battery_trace.get('charge_kw')
        discharge_arr = battery_trace.get('discharge_kw')
        charge_pv_arr = battery_trace.get('charge_from_pv_kw')
        charge_grid_arr = battery_trace.get('charge_from_grid_kw')
        discharge_load_arr = battery_trace.get('discharge_to_load_kw')
        discharge_grid_arr = battery_trace.get('discharge_to_grid_kw')
        pv_arr = battery_trace.get('pv_kw')
        load_arr = battery_trace.get('load_kw')

        n_rows = n_timesteps
        for arr in [soc_arr, charge_arr, discharge_arr]:
            if arr is not None:
                n_rows = min(n_rows, len(arr))

        if charge_pv_arr is None and charge_arr is not None and pv_arr is not None and load_arr is not None:
            surplus = np.maximum(pv_arr[:n_rows] - load_arr[:n_rows], 0.0)
            charge_pv_arr = np.minimum(charge_arr[:n_rows], surplus)
            charge_grid_arr = np.maximum(charge_arr[:n_rows] - charge_pv_arr, 0.0)
        if discharge_load_arr is None and discharge_arr is not None:
            discharge_load_arr = discharge_arr[:n_rows]
            discharge_grid_arr = np.zeros(n_rows)

        # Title
        ws_b['A1'] = "Praca Magazynu Energii - Dane Godzinowe"
        ws_b['A1'].font = Font(bold=True, size=14)
        ws_b.merge_cells('A1:S1')
        ws_b['A2'] = f"Magazyn: {bess_power_kw} kW / {bess_energy_kwh} kWh | Okres: {start_date} - {end_date}"
        ws_b['A2'].font = Font(italic=True, color="666666")
        ws_b.merge_cells('A2:S2')

        # Headers row 4
        # A=datetime B=weekday C=hour D=pv E=load F=pv_direct(formula)
        # G=soc H=soc%(formula) I=charge J=charge_pv K=charge_grid L=charge_kwh(formula)
        # M=discharge N=disch_load O=disch_grid P=disch_kwh(formula)
        # Q=bl_import R=pj_import S=reduction(formula)
        b_headers = [
            ('Data i godzina', 18, None),
            ('Dzień tyg.', 8, None),
            ('Godzina', 8, None),
            ('Produkcja PV\n[kW]', 14, pv_fill),
            ('Zużycie obiektu\n[kW]', 16, pv_fill),
            ('PV zużyte\nna miejscu [kW]', 16, formula_fill),  # FORMULA
            ('Stan magazynu\n[kWh]', 14, batt_fill),
            ('Stan magazynu\n[%]', 12, formula_fill),  # FORMULA
            ('Ładowanie\nmag. [kW]', 14, charge_fill),
            ('Ładowanie\nz PV [kW]', 14, charge_fill),
            ('Ładowanie\nz sieci [kW]', 14, charge_fill),
            ('Energia\nzaładowana [kWh]', 16, formula_fill),  # FORMULA
            ('Rozładowanie\nmag. [kW]', 14, discharge_fill),
            ('Rozład. na\nobiekt [kW]', 14, discharge_fill),
            ('Rozład. do\nsieci [kW]', 14, discharge_fill),
            ('Energia\nrozładowana [kWh]', 16, formula_fill),  # FORMULA
            ('Pobór z sieci\nbez mag. [kW]', 16, None),
            ('Pobór z sieci\nz mag. [kW]', 16, None),
            ('Zmniejszenie\npoboru [kW]', 16, formula_fill),  # FORMULA
        ]

        for ci, (h, w, fill) in enumerate(b_headers, 1):
            cell = ws_b.cell(row=4, column=ci, value=h)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws_b.column_dimensions[get_column_letter(ci)].width = w

        for i in range(n_rows):
            r = i + 5
            dt = time_index[i]

            ws_b.cell(row=r, column=1, value=dt.replace(tzinfo=None))
            ws_b.cell(row=r, column=2, value=['Pon', 'Wt', 'Śr', 'Czw', 'Pt', 'Sob', 'Nd'][dt.weekday()])
            ws_b.cell(row=r, column=3, value=dt.hour)

            # D: PV (static)
            if pv_arr is not None:
                ws_b.cell(row=r, column=4, value=round(float(pv_arr[i]), 1)).fill = pv_fill
            # E: Load (static)
            if load_arr is not None:
                ws_b.cell(row=r, column=5, value=round(float(load_arr[i]), 1)).fill = pv_fill

            # F: PV zużyte na miejscu = MIN(D, E) (FORMULA)
            c = ws_b.cell(row=r, column=6, value=f'=MIN(D{r},E{r})')
            c.fill = formula_fill; c.number_format = '0.0'

            # G: SOC [kWh] (static)
            if soc_arr is not None:
                ws_b.cell(row=r, column=7, value=round(float(soc_arr[i]), 1)).fill = batt_fill

            # H: SOC [%] = G / pojemność * 100 (FORMULA)
            c = ws_b.cell(row=r, column=8, value=f'=IF(G{r}="","",G{r}/{bess_energy_kwh}*100)')
            c.fill = formula_fill; c.number_format = '0.0'

            # I: Ładowanie [kW] (static)
            if charge_arr is not None:
                ws_b.cell(row=r, column=9, value=round(float(charge_arr[i]), 1)).fill = charge_fill
            # J: Ładowanie z PV (static)
            if charge_pv_arr is not None:
                ws_b.cell(row=r, column=10, value=round(float(charge_pv_arr[i]), 1)).fill = charge_fill
            # K: Ładowanie z sieci (static)
            if charge_grid_arr is not None:
                ws_b.cell(row=r, column=11, value=round(float(charge_grid_arr[i]), 1)).fill = charge_fill

            # L: Energia załadowana [kWh] = I * dt_hours (FORMULA)
            c = ws_b.cell(row=r, column=12, value=f'=I{r}*{dt_hours}')
            c.fill = formula_fill; c.number_format = '0.00'

            # M: Rozładowanie [kW] (static)
            if discharge_arr is not None:
                ws_b.cell(row=r, column=13, value=round(float(discharge_arr[i]), 1)).fill = discharge_fill
            # N: Rozładowanie na obiekt (static)
            if discharge_load_arr is not None:
                ws_b.cell(row=r, column=14, value=round(float(discharge_load_arr[i]), 1)).fill = discharge_fill
            # O: Rozładowanie do sieci (static)
            if discharge_grid_arr is not None:
                ws_b.cell(row=r, column=15, value=round(float(discharge_grid_arr[i]), 1)).fill = discharge_fill

            # P: Energia rozładowana [kWh] = M * dt_hours (FORMULA)
            c = ws_b.cell(row=r, column=16, value=f'=M{r}*{dt_hours}')
            c.fill = formula_fill; c.number_format = '0.00'

            # Q: Pobór bez mag. [kW] (static)
            if i < len(baseline_import_kw):
                ws_b.cell(row=r, column=17, value=round(float(baseline_import_kw[i]), 1))
            # R: Pobór z mag. [kW] (static)
            if i < len(project_import_kw):
                ws_b.cell(row=r, column=18, value=round(float(project_import_kw[i]), 1))

            # S: Zmniejszenie poboru = Q - R (FORMULA)
            c = ws_b.cell(row=r, column=19, value=f'=Q{r}-R{r}')
            c.fill = formula_fill; c.number_format = '0.0'; c.font = Font(bold=True)

        # Summary row 3 with FORMULAS
        last_b_row = n_rows + 4
        ws_b.cell(row=3, column=1, value="SUMY:").font = Font(bold=True)

        # SUM formulas for summary
        for col, letter in [(6, 'F'), (12, 'L'), (16, 'P'), (19, 'S')]:
            c = ws_b.cell(row=3, column=col, value=f'=SUM({letter}5:{letter}{last_b_row})')
            c.font = Font(bold=True); c.number_format = '#,##0.0'

        ws_b.freeze_panes = 'A5'

    # =====================================================================
    # SHEET: Arbitraż RDN — godzinowy rozkład cen, stanów i zysków
    # =====================================================================
    rdn_prices = kwargs.get('hourly_prices_pln_mwh', None)
    if (rdn_prices and len(rdn_prices) >= 100 and
            battery_trace and isinstance(battery_trace, dict)):
        _build_arbitrage_rdn_sheet(
            wb, battery_trace, rdn_prices, time_index, dt_hours,
            n_timesteps, bess_power_kw, bess_energy_kwh, start_date, end_date,
            hdr_fill, hdr_font, charge_fill, discharge_fill, batt_fill,
            formula_fill, savings_fill, pv_fill, nfmt,
            pv_arr=kwargs.get('pv_kw'),
            load_arr=kwargs.get('load_kw'),
        )

        # Sheet: Arbitraż All-In — same structure but with full network fee breakdown
        _build_arbitrage_allin_sheet(
            wb, battery_trace, rdn_prices, time_index, dt_hours,
            n_timesteps, bess_power_kw, bess_energy_kwh, start_date, end_date,
            tou_config, fixed_config, bl_mocowa, pj_mocowa,
            hdr_fill, hdr_font, charge_fill, discharge_fill, batt_fill,
            formula_fill, savings_fill, pv_fill, nfmt,
            pv_arr=kwargs.get('pv_kw'),
            load_arr=kwargs.get('load_kw'),
        )

    # =====================================================================
    # SHEET: Warianty S/M/L  (sizing variant comparison)
    # =====================================================================
    sizing_variants = kwargs.get('sizing_variants', None)
    if sizing_variants and len(sizing_variants) >= 1:
        ws_v = wb.create_sheet("Warianty S-M-L")
        _build_variants_sheet(ws_v, sizing_variants, hdr_fill, hdr_font, sec_fill, sec_font, nfmt)

    # =====================================================================
    # SHEET: Pełna siatka  (full grid search results)
    # =====================================================================
    grid_search_results = kwargs.get('grid_search_results', None)
    if grid_search_results and len(grid_search_results) >= 1:
        ws_g = wb.create_sheet("Pełna siatka doboru")
        _build_grid_search_sheet(ws_g, grid_search_results, hdr_fill, hdr_font, savings_fill, nfmt)

    # =====================================================================
    # SHEET: Cykl Życia (Lifecycle cashflow year by year)
    # =====================================================================
    if sizing_variants:
        # Find recommended variant with cashflow_timeseries
        rec_variant = None
        for sv in sizing_variants:
            fs = sv.get('finance_summary') if isinstance(sv, dict) else getattr(sv, 'finance_summary', None)
            if fs:
                cf = fs.get('cashflow_timeseries') if isinstance(fs, dict) else getattr(fs, 'cashflow_timeseries', None)
                if cf and len(cf) > 0:
                    if rec_variant is None or (isinstance(sv, dict) and sv.get('is_recommended')):
                        rec_variant = sv

        if rec_variant:
            fs = rec_variant.get('finance_summary') if isinstance(rec_variant, dict) else getattr(rec_variant, 'finance_summary', None)
            cf_data = (fs.get('cashflow_timeseries') if isinstance(fs, dict) else getattr(fs, 'cashflow_timeseries', None)) or []
            if cf_data:
                ws_lc = wb.create_sheet("Cykl Życia")
                _build_lifecycle_sheet(ws_lc, rec_variant, cf_data, hdr_fill, hdr_font, sec_fill, sec_font, nfmt, savings_fill)

    # =====================================================================
    # SHEET: Optimum EFC (lifecycle NPV vs cycle limit)
    # =====================================================================
    if sizing_variants:
        for sv in sizing_variants:
            fs2 = sv.get('finance_summary') if isinstance(sv, dict) else getattr(sv, 'finance_summary', None)
            if fs2:
                efc_sweep = fs2.get('efc_optimization_sweep') if isinstance(fs2, dict) else getattr(fs2, 'efc_optimization_sweep', None)
                if efc_sweep and len(efc_sweep) > 0:
                    ws_efc = wb.create_sheet("Optimum EFC")
                    ws_efc.sheet_properties.tabColor = "FF6F00"

                    ws_efc.merge_cells('A1:G1')
                    c = ws_efc.cell(row=1, column=1, value="OPTYMALIZACJA LIMITU CYKLI (EFC)")
                    c.font = Font(bold=True, size=14, color="FFFFFF")
                    c.fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")

                    ws_efc.cell(row=3, column=1, value="Więcej cykli = więcej zarobków ale krótszy EOL baterii.").font = Font(italic=True, size=10, color="666666")
                    ws_efc.cell(row=4, column=1, value="Optimum = limit cykli przy którym lifecycle NPV jest najwyższe.").font = Font(italic=True, size=10, color="666666")

                    efc_headers = ['Limit EFC/rok', 'Oszczędności/rok [PLN]', 'EOL [rok]', 'Lifecycle NPV [PLN]', 'PLN/EFC', 'Optymalny']
                    for j, h in enumerate(efc_headers, 1):
                        c = ws_efc.cell(row=6, column=j, value=h)
                        c.font = hdr_font
                        c.fill = hdr_fill
                        c.alignment = Alignment(horizontal='center')

                    for j, w in enumerate([14, 20, 10, 20, 12, 12], 1):
                        ws_efc.column_dimensions[chr(64 + j)].width = w

                    for i, pt in enumerate(efc_sweep):
                        row = 7 + i
                        ws_efc.cell(row=row, column=1, value=pt.get('efc_limit', 0)).alignment = Alignment(horizontal='center')
                        ws_efc.cell(row=row, column=2, value=round(pt.get('annual_savings_pln', 0), 0)).number_format = '#,##0'
                        ws_efc.cell(row=row, column=3, value=pt.get('eol_year', 0)).alignment = Alignment(horizontal='center')
                        ws_efc.cell(row=row, column=4, value=round(pt.get('lifecycle_npv_pln', 0), 0)).number_format = '#,##0'
                        ws_efc.cell(row=row, column=5, value=round(pt.get('pln_per_efc', 0), 1)).number_format = '#,##0.0'

                        is_opt = pt.get('is_optimal', False)
                        if is_opt:
                            ws_efc.cell(row=row, column=6, value="★ OPTIMUM").font = Font(bold=True, color="006600")
                            for j2 in range(1, 7):
                                ws_efc.cell(row=row, column=j2).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

                    break  # Only first variant with sweep

    # =====================================================================
    # Save
    # =====================================================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =============================================================================
# Helper: Warianty S/M/L sheet
# =============================================================================

def _build_variants_sheet(ws, variants, hdr_fill, hdr_font, sec_fill, sec_font, nfmt):
    """Build comparison sheet for S/M/L variants."""
    ws.sheet_properties.tabColor = "4CAF50"

    # Title
    ws.merge_cells('A1:H1')
    c = ws.cell(row=1, column=1, value="Porównanie wariantów doboru BESS (S / M / L)")
    c.font = Font(bold=True, size=14, color="1565C0")
    ws.cell(row=2, column=1, value="Porównanie pojemności z metrykami degradacji i budżetem cykli").font = Font(italic=True, color="666666")

    # Sort: Small, Medium, Large
    order = {'small': 0, 'medium': 1, 'large': 2, '1h': 0, '2h': 1, '4h': 2}
    variants_sorted = sorted(variants, key=lambda v: order.get(str(v.get('variant', v.get('variant_label', ''))).lower(), 99))

    # --- Section 1: Parametry techniczne ---
    row = 4
    ws.merge_cells(f'A{row}:H{row}')
    c = ws.cell(row=row, column=1, value="PARAMETRY TECHNICZNE I EKONOMICZNE")
    c.fill = hdr_fill; c.font = hdr_font

    row = 5
    param_labels = [
        ("Wariant", None),
        ("Czas trwania [h]", None),
        ("Moc [kW]", nfmt),
        ("Pojemność [kWh]", nfmt),
        ("", None),
        ("CAPEX [tys. PLN]", '#,##0.0'),
        ("Oszczędności/rok [PLN]", '#,##0'),
        ("NPV [tys. PLN]", '#,##0.0'),
        ("Payback [lat]", '0.0'),
        ("", None),
        ("Autokonsumpcja / redukcja importu [PLN]", '#,##0'),
        ("Peak shaving (opłata za moc) [PLN]", '#,##0'),
        ("Opłata mocowa PL [PLN]", '#,##0'),
        ("Arbitraż ToU [PLN]", '#,##0'),
        ("Koszt degradacji [PLN]", '#,##0'),
        ("", None),
        ("Throughput [MWh/rok]", '#,##0.0'),
        ("EFC łącznie [cykli/rok]", '#,##0'),
        ("Ładowanie z PV [%]", '0.0'),
        ("Ładowanie z sieci [MWh]", '#,##0.0'),
        ("Żywotność [lat]", '0'),
        ("Status degradacji", None),
    ]

    # Write row labels
    for i, (label, _) in enumerate(param_labels):
        r = row + i
        if label == "":
            continue
        c = ws.cell(row=r, column=1, value=label)
        if label.startswith("CAPEX") or label.startswith("Oszczędności") or label.startswith("NPV") or label.startswith("Payback"):
            c.font = Font(bold=True)

    # Write variant columns
    for vi, v in enumerate(variants_sorted):
        col = 2 + vi
        ws.cell(row=4, column=col).fill = hdr_fill

        data_rows = [
            v.get('variant_label', v.get('variant', f'Wariant {vi+1}')),
            v.get('duration_h', ''),
            v.get('power_kw', 0),
            v.get('energy_kwh', 0),
            None,  # spacer
            round(v.get('capex_pln', 0) / 1000, 1),
            round(v.get('annual_savings_pln', 0)),
            round(v.get('npv_pln', 0) / 1000, 1),
            v.get('simple_payback_years', v.get('payback_years', 0)),
            None,  # spacer
            round(v.get('energy_savings_pln', 0)),
            round(v.get('demand_charge_savings_pln', v.get('peak_shaving_savings_pln', 0))),
            round(v.get('capacity_fee_savings_pln', 0)),
            round(v.get('arbitrage_savings_pln', 0)),
            round(v.get('degradation_cost_pln', 0)),
            None,  # spacer
            v.get('throughput_mwh_yr', 0),
            v.get('efc_total', 0),
            v.get('pv_charge_pct', v.get('self_consumption_pct', 0)),
            v.get('grid_charge_mwh', 0),
            v.get('lifetime_years', 0),
            v.get('degradation_status', ''),
        ]

        for i, val in enumerate(data_rows):
            r = row + i
            if val is None:
                continue
            c = ws.cell(row=r, column=col, value=val)
            _, fmt = param_labels[i]
            if fmt:
                c.number_format = fmt

    # Column widths
    ws.column_dimensions['A'].width = 42
    for vi in range(len(variants_sorted)):
        ws.column_dimensions[get_column_letter(2 + vi)].width = 20

    # Highlight recommended variant
    for vi, v in enumerate(variants_sorted):
        if v.get('is_recommended'):
            col = 2 + vi
            rec_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            for r in range(row, row + len(param_labels)):
                ws.cell(row=r, column=col).fill = rec_fill
            ws.cell(row=row + len(param_labels), column=col, value="REKOMENDOWANY").font = Font(bold=True, color="2E7D32")


# =============================================================================
# Helper: Pełna siatka doboru sheet
# =============================================================================

def _build_lifecycle_sheet(ws, variant, cashflow_data, hdr_fill, hdr_font, sec_fill, sec_font, nfmt, savings_fill):
    """Build lifecycle cashflow year-by-year sheet."""
    from openpyxl.styles import Alignment, Border, Side, numbers
    ws.sheet_properties.tabColor = "2E7D32"

    thin = Side(style='thin', color='999999')
    border = Border(bottom=thin)

    # Get variant info
    v = variant if isinstance(variant, dict) else {}
    label = v.get('variant_label', 'BESS')
    power_kw = v.get('power_kw', 0)
    energy_kwh = v.get('energy_kwh', 0)
    capex = v.get('capex_pln', 0)
    fs = v.get('finance_summary', {}) or {}
    irr = fs.get('irr_pct')
    horizon = fs.get('horizon_years', len(cashflow_data) - 1)
    dr = fs.get('discount_rate', 0.10)

    # Title
    ws.merge_cells('A1:H1')
    c = ws.cell(row=1, column=1, value=f"CYKL ŻYCIA PROJEKTU — {label}")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    ws.row_dimensions[1].height = 30

    # Parameters section
    params = [
        ('Moc [kW]', round(power_kw, 1)),
        ('Pojemność [kWh]', round(energy_kwh, 1)),
        ('CAPEX [tys. PLN]', round(capex / 1000, 1)),
        ('Horyzont [lat]', horizon),
        ('Stopa dyskontowa', dr),
        ('IRR', irr / 100 if irr else None),
    ]
    for i, (name, val) in enumerate(params):
        ws.cell(row=3 + i, column=1, value=name).font = Font(bold=True, size=10)
        c = ws.cell(row=3 + i, column=2, value=val)
        if 'dyskontowa' in name or name == 'IRR':
            c.number_format = '0.0%'
        elif 'tys' in name:
            c.number_format = '#,##0.0'
        else:
            c.number_format = '#,##0.0'

    # Cashflow table header
    header_row = 11
    headers = [
        'Rok', 'Oszczędności [PLN]', 'OPEX [PLN]', 'Net Cashflow [PLN]',
        'Skumulowany [PLN]', 'Zdyskontowany [PLN]', 'SoH [%]'
    ]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Column widths
    widths = [8, 18, 14, 18, 18, 18, 10]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + j)].width = w

    # Data rows
    for i, cf in enumerate(cashflow_data):
        row = header_row + 1 + i
        if isinstance(cf, dict):
            year = cf.get('year', i)
            savings = cf.get('savings_pln', 0)
            opex = cf.get('opex_pln', 0)
            net = cf.get('net_cashflow_pln', 0)
            cumul = cf.get('cumulative_cashflow_pln', 0)
            disc = cf.get('discounted_cashflow_pln', 0)
        else:
            year = getattr(cf, 'year', i)
            savings = getattr(cf, 'savings_pln', 0)
            opex = getattr(cf, 'opex_pln', 0)
            net = getattr(cf, 'net_cashflow_pln', 0)
            cumul = getattr(cf, 'cumulative_cashflow_pln', 0)
            disc = getattr(cf, 'discounted_cashflow_pln', 0)

        ws.cell(row=row, column=1, value=year).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=round(savings, 0)).number_format = '#,##0'
        ws.cell(row=row, column=3, value=round(opex, 0)).number_format = '#,##0'

        c_net = ws.cell(row=row, column=4, value=round(net, 0))
        c_net.number_format = '#,##0'
        if net < 0:
            c_net.font = Font(color="CC0000")
        else:
            c_net.font = Font(color="006600")

        c_cum = ws.cell(row=row, column=5, value=round(cumul, 0))
        c_cum.number_format = '#,##0'
        if cumul < 0:
            c_cum.font = Font(color="CC0000")
        else:
            c_cum.font = Font(color="006600")
            c_cum.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

        ws.cell(row=row, column=6, value=round(disc, 0)).number_format = '#,##0'

        # SoH from backend (precise combined calendar + cycle model)
        soh_pct = cf.get('soh_pct') if isinstance(cf, dict) else getattr(cf, 'soh_pct', None)
        is_eol = cf.get('is_eol') if isinstance(cf, dict) else getattr(cf, 'is_eol', None)
        if soh_pct is not None:
            c_soh = ws.cell(row=row, column=7, value=soh_pct / 100.0)
            c_soh.number_format = '0.0%'
            if is_eol:
                c_soh.font = Font(color="CC0000", bold=True)
                c_soh.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        elif year == 0:
            ws.cell(row=row, column=7, value=1.0).number_format = '0.0%'

        # Thin border
        for j in range(1, 8):
            ws.cell(row=row, column=j).border = border

    # NPV summary row
    last_data_row = header_row + len(cashflow_data)
    summary_row = last_data_row + 2
    ws.cell(row=summary_row, column=1, value='NPV').font = Font(bold=True, size=12)
    npv_val = fs.get('npv_pln', 0)
    c = ws.cell(row=summary_row, column=4, value=round(npv_val, 0))
    c.number_format = '#,##0'
    c.font = Font(bold=True, size=12, color="006600" if npv_val >= 0 else "CC0000")

    ws.cell(row=summary_row + 1, column=1, value='IRR').font = Font(bold=True, size=12)
    if irr is not None:
        ws.cell(row=summary_row + 1, column=4, value=irr / 100).number_format = '0.0%'
        ws.cell(row=summary_row + 1, column=4).font = Font(bold=True, size=12)

    # Find EOL year dynamically
    eol_year = None
    for cf in cashflow_data:
        is_eol = cf.get('is_eol') if isinstance(cf, dict) else getattr(cf, 'is_eol', None)
        if is_eol:
            eol_year = cf.get('year') if isinstance(cf, dict) else getattr(cf, 'year', None)
            break

    eol_row = summary_row + 3
    ws.cell(row=eol_row, column=1, value='EOL baterii (SoH < 70%)').font = Font(bold=True, size=11, color="CC0000")
    if eol_year:
        ws.cell(row=eol_row, column=4, value=f"Rok {eol_year}").font = Font(bold=True, size=11, color="CC0000")
    else:
        ws.cell(row=eol_row, column=4, value=f"Poza horyzontem ({horizon} lat)").font = Font(size=11, color="006600")


def _build_grid_search_sheet(ws, grid_results, hdr_fill, hdr_font, savings_fill, nfmt):
    """Build full grid search results table."""
    ws.sheet_properties.tabColor = "FF9800"

    # Title
    ws.merge_cells('A1:R1')
    c = ws.cell(row=1, column=1, value="Pełna siatka przeszukanych konfiguracji BESS")
    c.font = Font(bold=True, size=14, color="1565C0")
    ws.cell(row=2, column=1, value=f"Łącznie {len(grid_results)} konfiguracji, posortowane wg NPV/kWh malejąco").font = Font(italic=True, color="666666")

    # Headers
    headers = [
        ("Lp.", 5),
        ("Moc [kW]", 12),
        ("Energia [kWh]", 14),
        ("Czas [h]", 8),
        ("CAPEX [tys. PLN]", 16),
        ("Autokonsumpcja [PLN]", 20),
        ("Arbitraż [PLN]", 14),
        ("Peak shaving [PLN]", 16),
        ("Opłata mocowa [PLN]", 18),
        ("Usł. sieciowe [PLN]", 16),
        ("RAZEM [PLN/rok]", 16),
        ("NPV [tys. PLN]", 14),
        ("NPV/kWh [PLN]", 14),
        ("Payback [lat]", 12),
        ("Żywotność [lat]", 14),
        ("EFC [cykli/rok]", 14),
        ("Kara EFC [tys.]", 14),
        ("Autokons. [%]", 12),
        ("Peak red. [%]", 12),
    ]

    row = 4
    for ci, (hdr, width) in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=hdr)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows - sort by npv_per_kwh descending
    sorted_results = sorted(grid_results, key=lambda x: x.get('npv_per_kwh', 0), reverse=True)

    positive_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    negative_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for idx, g in enumerate(sorted_results):
        r = row + 1 + idx
        npv = g.get('npv_pln', 0)
        row_fill = positive_fill if npv > 0 else (negative_fill if npv < 0 else None)

        dur_h = g.get('duration_h', 0)
        dur_str = f"{int(dur_h)}h" if dur_h == int(dur_h) else f"{dur_h}h"

        energy_sav = g.get('energy_savings_pln', 0)
        arb_sav = g.get('arbitrage_savings_pln', 0)
        peak_sav = g.get('peak_shaving_savings_pln', g.get('demand_charge_savings_pln', 0))
        cap_sav = g.get('capacity_fee_savings_pln', 0)
        ancillary = g.get('ancillary_revenue_pln', 0)
        annual = g.get('annual_savings_pln', 0)
        degrad_cost = g.get('degradation_cost_pln', 0)

        vals = [
            (idx + 1, '0'),
            (round(g.get('power_kw', 0)), '#,##0'),
            (round(g.get('energy_kwh', 0)), '#,##0'),
            (dur_str, None),
            (round(g.get('capex_pln', 0) / 1000, 0), '#,##0'),
            (round(energy_sav), '#,##0'),
            (round(arb_sav), '#,##0'),
            (round(peak_sav), '#,##0'),
            (round(cap_sav), '#,##0'),
            (round(ancillary), '#,##0'),
            (round(annual), '#,##0'),
            (round(npv / 1000, 0), '#,##0'),
            (round(g.get('npv_per_kwh', 0)), '#,##0'),
            (round(g.get('payback_years', 0), 1), '0.0'),
            (round(g.get('lifetime_years', 0)), '0'),
            (round(g.get('efc_total', 0)), '#,##0'),
            (round(g.get('efc_penalty_pln', 0) / 1000, 0) if g.get('efc_penalty_pln', 0) else '—', '#,##0'),
            (round(g.get('self_consumption_pct', 0), 1), '0.0'),
            (round(g.get('peak_reduction_pct', 0), 1), '0.0'),
        ]

        for ci, (val, fmt) in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            if fmt:
                c.number_format = fmt
            if row_fill:
                c.fill = row_fill

    # --- Best configurations section ---
    best_row = row + 1 + len(sorted_results) + 2
    ws.merge_cells(f'A{best_row}:R{best_row}')
    c = ws.cell(row=best_row, column=1, value="NAJLEPSZE KONFIGURACJE")
    c.font = Font(bold=True, size=13, color="1565C0")
    c.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

    criteria = [
        ("Najlepsza wartość (NPV/kWh)", lambda x: x.get('npv_per_kwh', -999999)),
        ("Najszybszy zwrot (Payback)", lambda x: -x.get('payback_years', 999) if x.get('npv_pln', 0) > 0 else -999),
        ("Najwyższe oszczędności roczne", lambda x: x.get('annual_savings_pln', 0)),
        ("Najwyższe NPV", lambda x: x.get('npv_pln', -999999)),
    ]

    cr = best_row + 1
    for crit_name, key_fn in criteria:
        ws.merge_cells(f'A{cr}:R{cr}')
        c = ws.cell(row=cr, column=1, value=crit_name)
        c.font = Font(bold=True, size=11, color="E65100")
        cr += 1

        # Headers for this section
        mini_hdrs = ["Moc [kW]", "Energia [kWh]", "Czas", "CAPEX [tys.]", "Autokon. [PLN]", "Arbitraż [PLN]",
                     "Peak [PLN]", "Op. moc. [PLN]", "RAZEM/rok [PLN]", "NPV [tys.]", "NPV/kWh", "Payback [lat]",
                     "EFC/rok", "Autokons. [%]"]
        for ci, h in enumerate(mini_hdrs, 1):
            c = ws.cell(row=cr, column=ci, value=h)
            c.font = Font(bold=True, size=9); c.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        cr += 1

        best = sorted(sorted_results, key=key_fn, reverse=True)[:3]
        for bi, g in enumerate(best):
            energy_sav = g.get('energy_savings_pln', 0)
            arb_sav = g.get('arbitrage_savings_pln', 0)
            peak_sav = g.get('peak_shaving_savings_pln', g.get('demand_charge_savings_pln', 0))
            cap_sav = g.get('capacity_fee_savings_pln', 0)
            annual = g.get('annual_savings_pln', 0)
            dur_h = g.get('duration_h', 0)
            dur_str = f"{int(dur_h)}h" if dur_h == int(dur_h) else f"{dur_h}h"

            mini_vals = [
                round(g.get('power_kw', 0)),
                round(g.get('energy_kwh', 0)),
                dur_str,
                round(g.get('capex_pln', 0) / 1000),
                round(energy_sav),
                round(arb_sav),
                round(peak_sav),
                round(cap_sav),
                round(annual),
                round(g.get('npv_pln', 0) / 1000),
                round(g.get('npv_per_kwh', 0)),
                round(g.get('payback_years', 0), 1),
                round(g.get('efc_total', 0)),
                round(g.get('self_consumption_pct', 0), 1),
            ]
            medal = ["🥇", "🥈", "🥉"][bi] if bi < 3 else ""
            ws.cell(row=cr, column=1, value=f"{medal} {mini_vals[0]}").number_format = '#,##0'
            for ci, val in enumerate(mini_vals[1:], 2):
                c = ws.cell(row=cr, column=ci, value=val)
                if isinstance(val, (int, float)):
                    c.number_format = '#,##0' if isinstance(val, int) else '0.0'
            if bi == 0:
                gold_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                for ci in range(1, 15):
                    ws.cell(row=cr, column=ci).fill = gold_fill
            cr += 1
        cr += 1  # gap between criteria

    ws.freeze_panes = 'A5'


# =============================================================================
# Helper: Arbitraż All-In — pełne rozbicie kosztów sieciowych
# =============================================================================

def _build_arbitrage_allin_sheet(
    wb, battery_trace, rdn_prices, time_index, dt_hours,
    n_timesteps, bess_power_kw, bess_energy_kwh, start_date, end_date,
    tou_config, fixed_config, bl_mocowa, pj_mocowa,
    hdr_fill, hdr_font, charge_fill, discharge_fill, batt_fill,
    formula_fill, savings_fill, pv_fill, nfmt,
    pv_arr=None, load_arr=None,
):
    """Build 'Arbitraż All-In' sheet — full network fee breakdown per hour."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Arbitraż All-In")

    charge_arr = battery_trace.get('charge_kw', [])
    discharge_arr = battery_trace.get('discharge_kw', [])
    soc_arr = battery_trace.get('soc_kwh', [])

    n_rows = min(n_timesteps, len(rdn_prices))
    for arr in [charge_arr, discharge_arr, soc_arr]:
        if arr is not None:
            n_rows = min(n_rows, len(arr))

    rdn = np.array(rdn_prices[:n_rows], dtype=float)

    # Fills
    rdn_cheap_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    rdn_mid_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    rdn_exp_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    fee_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")      # indigo light
    allin_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")     # pink light
    profit_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
    loss_fill = PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid")
    summary_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")

    p25 = float(np.percentile(rdn, 25))
    p75 = float(np.percentile(rdn, 75))

    weekdays_pl = ['Pon', 'Wt', 'Śr', 'Czw', 'Pt', 'Sob', 'Nd']

    # Fee rates (PLN/MWh)
    quality = fixed_config.quality_fee
    oze = fixed_config.oze_fee
    cogen = fixed_config.cogeneration_fee
    excise = fixed_config.excise_tax

    # Title
    ws['A1'] = "Arbitraż All-In — Pełne koszty sieciowe w rozbiciu godzinowym"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:V1')

    total_fees = quality + oze + cogen + excise
    ws['A2'] = (
        f"Magazyn: {bess_power_kw:.0f} kW / {bess_energy_kwh:.0f} kWh | "
        f"Opłaty stałe: jakość={quality:.0f}, OZE={oze:.0f}, kog.={cogen:.0f}, "
        f"akcyza={excise:.0f} PLN/MWh | SOM={fixed_config.capacity_fee_som:.4f} PLN/kWh"
    )
    ws['A2'].font = Font(italic=True, color="666666")
    ws.merge_cells('A2:V2')

    # Headers row 4
    # A=timestamp B=weekday C=hour D=RDN E=Dystrybucja F=Jakościowa G=OZE H=Kogeneracja
    # I=Akcyza J=Mocowa K=All-in [PLN/MWh] L=SoC M=Charge N=Discharge O=Stan
    # P=Koszt ładow. all-in Q=Wartość rozład. all-in R=Zysk netto S=Zysk dzień T=Zysk miesiąc U=Zysk rok
    headers = [
        ('Data i godzina', 18),
        ('Dzień', 8),
        ('Godz', 6),
        ('Cena RDN\n[PLN/MWh]', 12),
        ('Dystrybucja\n[PLN/MWh]', 12),
        ('Jakościowa\n[PLN/MWh]', 11),
        ('OZE\n[PLN/MWh]', 10),
        ('Kogeneracja\n[PLN/MWh]', 11),
        ('Akcyza\n[PLN/MWh]', 10),
        ('Mocowa\n[PLN/godz]', 11),
        ('ALL-IN\n[PLN/MWh]', 13),
        ('SoC\n[kWh]', 10),
        ('Ładowanie\n[kW]', 12),
        ('Rozładowanie\n[kW]', 12),
        ('Stan', 14),
        ('Koszt ładow.\nALL-IN [PLN]', 15),
        ('Wartość rozład.\nALL-IN [PLN]', 15),
        ('Zysk netto\n[PLN]', 13),
        ('Zysk dzień\n[PLN]', 13),
        ('Zysk miesiąc\n[PLN]', 13),
        ('Zysk rok\n[PLN]', 13),
    ]

    for ci, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    for i in range(n_rows):
        r = i + 5
        dt = time_index[i]
        hour = dt.hour
        price = float(rdn[i])
        ch = float(charge_arr[i]) if i < len(charge_arr) else 0.0
        dis = float(discharge_arr[i]) if i < len(discharge_arr) else 0.0
        soc = float(soc_arr[i]) if i < len(soc_arr) else 0.0

        # Distribution rate for this hour
        dist_rate = get_distribution_rate(dt, tou_config, fixed_config)

        # Capacity fee for this hour (from pre-calculated arrays, per kWh imported)
        # bl_mocowa is PLN for baseline import at this hour
        # We want PLN/MWh equivalent — but mocowa is allocated proportionally
        # Use the project mocowa directly as PLN value for this hour
        mocowa_pln = float(pj_mocowa[i]) if i < len(pj_mocowa) else 0.0

        # All-in price (PLN/MWh) = RDN + distribution + fees
        allin = price + dist_rate + quality + oze + cogen + excise
        # Note: mocowa is per-hour absolute PLN, not per MWh rate

        # A: Timestamp
        ws.cell(row=r, column=1, value=dt.replace(tzinfo=None))
        # B: Weekday
        ws.cell(row=r, column=2, value=weekdays_pl[dt.weekday()])
        # C: Hour
        ws.cell(row=r, column=3, value=hour)

        # D: RDN price
        c = ws.cell(row=r, column=4, value=round(price, 1))
        c.number_format = '0.0'
        if price <= p25:
            c.fill = rdn_cheap_fill
        elif price >= p75:
            c.fill = rdn_exp_fill
        else:
            c.fill = rdn_mid_fill

        # E: Distribution
        ws.cell(row=r, column=5, value=round(dist_rate, 1)).fill = fee_fill
        # F: Quality
        ws.cell(row=r, column=6, value=round(quality, 1)).fill = fee_fill
        # G: OZE
        ws.cell(row=r, column=7, value=round(oze, 1)).fill = fee_fill
        # H: Cogeneration
        ws.cell(row=r, column=8, value=round(cogen, 1)).fill = fee_fill
        # I: Excise
        ws.cell(row=r, column=9, value=round(excise, 1)).fill = fee_fill
        # J: Capacity fee (PLN absolute for this hour)
        c = ws.cell(row=r, column=10, value=round(mocowa_pln, 4))
        c.number_format = '0.0000'
        c.fill = fee_fill

        # K: All-in (formula: D + E + F + G + H + I, mocowa separate)
        c = ws.cell(row=r, column=11, value=f'=D{r}+E{r}+F{r}+G{r}+H{r}+I{r}')
        c.number_format = '0.0'
        c.fill = allin_fill
        c.font = Font(bold=True)

        # L: SoC
        ws.cell(row=r, column=12, value=round(soc, 1)).fill = batt_fill

        # M: Charge
        c = ws.cell(row=r, column=13, value=round(ch, 1))
        if ch > 0.1:
            c.fill = charge_fill
            c.font = Font(bold=True)

        # N: Discharge
        c = ws.cell(row=r, column=14, value=round(dis, 1))
        if dis > 0.1:
            c.fill = discharge_fill
            c.font = Font(bold=True)

        # O: Status
        if ch > 0.1:
            c = ws.cell(row=r, column=15, value="ŁADOWANIE")
            c.fill = charge_fill
            c.font = Font(bold=True, color="1565C0")
        elif dis > 0.1:
            c = ws.cell(row=r, column=15, value="ROZŁADOWANIE")
            c.fill = discharge_fill
            c.font = Font(bold=True, color="E65100")
        else:
            ws.cell(row=r, column=15, value="—").alignment = Alignment(horizontal='center')

        # P: Koszt ładowania ALL-IN = charge_kw * dt * all-in/1000 + mocowa
        #    When battery charges from PV, load imports from grid at all-in price
        c = ws.cell(row=r, column=16, value=f'=M{r}*{dt_hours}*K{r}/1000+IF(M{r}>0,J{r},0)')
        c.number_format = '0.00'
        if ch > 0.1:
            c.fill = charge_fill

        # Q: Wartość rozładowania ALL-IN = discharge_kw * dt * all-in/1000 + mocowa saved
        #    When battery discharges, load avoids grid import at all-in price
        c = ws.cell(row=r, column=17, value=f'=N{r}*{dt_hours}*K{r}/1000+IF(N{r}>0,J{r},0)')
        c.number_format = '0.00'
        if dis > 0.1:
            c.fill = discharge_fill

        # R: Zysk netto = Q - P
        c = ws.cell(row=r, column=18, value=f'=Q{r}-P{r}')
        c.number_format = '0.00'
        c.fill = formula_fill

        # S: Zysk dzień (running per day)
        if i == 0 or time_index[i].date() != time_index[i - 1].date():
            c = ws.cell(row=r, column=19, value=f'=R{r}')
        else:
            c = ws.cell(row=r, column=19, value=f'=S{r-1}+R{r}')
        c.number_format = '0.00'
        c.fill = formula_fill

        # T: Zysk miesiąc
        if i == 0 or time_index[i].month != time_index[i - 1].month:
            c = ws.cell(row=r, column=20, value=f'=R{r}')
        else:
            c = ws.cell(row=r, column=20, value=f'=T{r-1}+R{r}')
        c.number_format = '0.00'
        c.fill = formula_fill

        # U: Zysk rok
        if i == 0:
            c = ws.cell(row=r, column=21, value=f'=R{r}')
        else:
            c = ws.cell(row=r, column=21, value=f'=U{r-1}+R{r}')
        c.number_format = '0.00'
        c.fill = savings_fill
        c.font = Font(bold=True)

        # Day separator
        if hour == 23:
            day_border = Border(bottom=Side(style='medium', color='999999'))
            for col in range(1, 22):
                ws.cell(row=r, column=col).border = day_border

    last_row = n_rows + 4

    # Summary row 3
    ws.cell(row=3, column=1, value="PODSUMOWANIE:").font = Font(bold=True, size=11)
    ws.cell(row=3, column=15, value="SUMY:").font = Font(bold=True)

    for col, letter in [(16, 'P'), (17, 'Q'), (18, 'R')]:
        c = ws.cell(row=3, column=col, value=f'=SUM({letter}5:{letter}{last_row})')
        c.font = Font(bold=True, size=11)
        c.number_format = '#,##0'
        c.fill = summary_fill if col < 18 else savings_fill

    # Final yearly profit in bold green
    ws.cell(row=3, column=18).font = Font(bold=True, size=12, color="1B5E20")

    # Monthly summary below data
    sr = last_row + 3
    ws.cell(row=sr, column=1, value="PODSUMOWANIE MIESIĘCZNE — ALL-IN").font = Font(bold=True, size=13)
    ws.merge_cells(f'A{sr}:H{sr}')
    sr += 1

    month_headers = [
        ('Miesiąc', 14), ('Ładow. [kWh]', 14), ('Rozład. [kWh]', 14),
        ('Koszt ład.\nall-in [PLN]', 15), ('Wartość rozł.\nall-in [PLN]', 15),
        ('ZYSK\n[PLN]', 14), ('Zysk RDN\nonly [PLN]', 14),
        ('Zysk z opłat\nsiec. [PLN]', 14),
    ]
    for ci, (h, w) in enumerate(month_headers, 1):
        cell = ws.cell(row=sr, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = max(
            ws.column_dimensions[get_column_letter(ci)].width or 0, w
        )
    sr += 1

    from collections import defaultdict
    monthly = defaultdict(lambda: {
        'charge_kwh': 0, 'discharge_kwh': 0,
        'charge_cost_allin': 0, 'discharge_value_allin': 0,
        'charge_cost_rdn': 0, 'discharge_value_rdn': 0,
    })

    for i in range(n_rows):
        m = time_index[i].month
        ch = float(charge_arr[i]) if i < len(charge_arr) else 0
        dis = float(discharge_arr[i]) if i < len(discharge_arr) else 0
        price = float(rdn[i])
        dist_rate = get_distribution_rate(time_index[i], tou_config, fixed_config)
        mocowa_pln = float(pj_mocowa[i]) if i < len(pj_mocowa) else 0.0
        allin = price + dist_rate + quality + oze + cogen + excise

        if ch > 0.1:
            energy_kwh = ch * dt_hours
            monthly[m]['charge_kwh'] += energy_kwh
            monthly[m]['charge_cost_allin'] += energy_kwh * allin / 1000 + mocowa_pln
            monthly[m]['charge_cost_rdn'] += energy_kwh * price / 1000
        if dis > 0.1:
            energy_kwh = dis * dt_hours
            monthly[m]['discharge_kwh'] += energy_kwh
            monthly[m]['discharge_value_allin'] += energy_kwh * allin / 1000 + mocowa_pln
            monthly[m]['discharge_value_rdn'] += energy_kwh * price / 1000

    month_names = ['', 'Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec',
                   'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']
    monthly_start = sr
    for m in sorted(monthly.keys()):
        mm = monthly[m]
        profit_allin = mm['discharge_value_allin'] - mm['charge_cost_allin']
        profit_rdn = mm['discharge_value_rdn'] - mm['charge_cost_rdn']
        profit_fees = profit_allin - profit_rdn

        ws.cell(row=sr, column=1, value=month_names[m]).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=round(mm['charge_kwh'], 0)).number_format = '#,##0'
        ws.cell(row=sr, column=3, value=round(mm['discharge_kwh'], 0)).number_format = '#,##0'
        ws.cell(row=sr, column=4, value=round(mm['charge_cost_allin'], 0)).number_format = '#,##0'
        ws.cell(row=sr, column=5, value=round(mm['discharge_value_allin'], 0)).number_format = '#,##0'

        c = ws.cell(row=sr, column=6, value=round(profit_allin, 0))
        c.number_format = '#,##0'
        c.font = Font(bold=True, size=11)
        c.fill = profit_fill if profit_allin >= 0 else loss_fill

        ws.cell(row=sr, column=7, value=round(profit_rdn, 0)).number_format = '#,##0'

        c = ws.cell(row=sr, column=8, value=round(profit_fees, 0))
        c.number_format = '#,##0'
        c.font = Font(bold=True, color="1565C0")

        sr += 1

    # Yearly totals
    c = ws.cell(row=sr, column=1, value="SUMA ROCZNA")
    c.font = Font(bold=True, size=12)
    for col in range(2, 9):
        letter = get_column_letter(col)
        c = ws.cell(row=sr, column=col, value=f'=SUM({letter}{monthly_start}:{letter}{sr-1})')
        c.font = Font(bold=True, size=12)
        c.number_format = '#,##0'
        c.fill = savings_fill

    ws.freeze_panes = 'A5'


# =============================================================================
# Helper: Arbitraż RDN sheet — hourly prices, battery state, profit analysis
# =============================================================================

def _build_arbitrage_rdn_sheet(
    wb, battery_trace, rdn_prices, time_index, dt_hours,
    n_timesteps, bess_power_kw, bess_energy_kwh, start_date, end_date,
    hdr_fill, hdr_font, charge_fill, discharge_fill, batt_fill,
    formula_fill, savings_fill, pv_fill, nfmt,
    pv_arr=None, load_arr=None,
):
    """Build 'Arbitraż RDN' sheet with hourly price/battery/profit analysis."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Arbitraż RDN")

    soc_arr = battery_trace.get('soc_kwh', [])
    charge_arr = battery_trace.get('charge_kw', [])
    discharge_arr = battery_trace.get('discharge_kw', [])
    charge_pv_arr = battery_trace.get('charge_from_pv_kw')
    charge_grid_arr = battery_trace.get('charge_from_grid_kw')

    n_rows = min(n_timesteps, len(rdn_prices))
    for arr in [soc_arr, charge_arr, discharge_arr]:
        if arr is not None:
            n_rows = min(n_rows, len(arr))

    rdn = np.array(rdn_prices[:n_rows], dtype=float)

    # --- Fills ---
    rdn_cheap_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")   # green — cheap
    rdn_mid_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")     # yellow — mid
    rdn_exp_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")     # red — expensive
    profit_fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")      # green — profit
    loss_fill = PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid")        # red — loss
    summary_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")     # light blue

    # Price percentiles for coloring
    p25 = float(np.percentile(rdn, 25))
    p75 = float(np.percentile(rdn, 75))

    # ===== Title =====
    ws['A1'] = "Arbitraż RDN — Analiza godzinowa cen i pracy magazynu"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:R1')

    ws['A2'] = (f"Magazyn: {bess_power_kw:.0f} kW / {bess_energy_kwh:.0f} kWh | "
                f"Okres: {start_date} - {end_date} | "
                f"Ceny RDN: min={rdn.min():.0f}, avg={rdn.mean():.0f}, max={rdn.max():.0f} PLN/MWh")
    ws['A2'].font = Font(italic=True, color="666666")
    ws.merge_cells('A2:R2')

    # ===== Summary row 3 (formulas filled after data) =====

    # ===== Headers row 4 =====
    # A=timestamp  B=weekday  C=hour  D=RDN price
    # E=PV  F=Load  G=SoC  H=SoC%
    # I=Charge  J=Discharge  K=Stan (charging/discharging/idle)
    # L=Koszt ładowania (PLN)  M=Wartość rozładowania (PLN)
    # N=Zysk godzinowy  O=Zysk skumulowany dzień  P=Zysk skumul. miesiąc  Q=Zysk skumul. rok
    headers = [
        ('Data i godzina', 18),
        ('Dzień', 8),
        ('Godz', 6),
        ('Cena RDN\n[PLN/MWh]', 13),
        ('PV\n[kW]', 10),
        ('Zużycie\n[kW]', 10),
        ('SoC\n[kWh]', 10),
        ('SoC\n[%]', 8),
        ('Ładowanie\n[kW]', 12),
        ('Rozładowanie\n[kW]', 12),
        ('Stan', 14),
        ('Koszt ładowania\n[PLN]', 16),
        ('Wartość rozład.\n[PLN]', 16),
        ('Zysk netto\ngodzina [PLN]', 14),
        ('Zysk skumul.\ndzień [PLN]', 14),
        ('Zysk skumul.\nmiesiąc [PLN]', 14),
        ('Zysk skumul.\nrok [PLN]', 14),
    ]

    for ci, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ===== Data rows (starting at row 5) =====
    thin_border = Border(
        bottom=Side(style='thin', color='CCCCCC'),
    )

    weekdays_pl = ['Pon', 'Wt', 'Śr', 'Czw', 'Pt', 'Sob', 'Nd']

    for i in range(n_rows):
        r = i + 5
        dt = time_index[i]
        hour = dt.hour
        price = float(rdn[i])
        ch = float(charge_arr[i]) if charge_arr is not None and i < len(charge_arr) else 0.0
        dis = float(discharge_arr[i]) if discharge_arr is not None and i < len(discharge_arr) else 0.0
        soc = float(soc_arr[i]) if soc_arr is not None and i < len(soc_arr) else 0.0

        # A: Timestamp
        ws.cell(row=r, column=1, value=dt.replace(tzinfo=None))

        # B: Weekday
        ws.cell(row=r, column=2, value=weekdays_pl[dt.weekday()])

        # C: Hour
        ws.cell(row=r, column=3, value=hour)

        # D: RDN price with conditional color
        c = ws.cell(row=r, column=4, value=round(price, 1))
        c.number_format = '0.0'
        if price <= p25:
            c.fill = rdn_cheap_fill
        elif price >= p75:
            c.fill = rdn_exp_fill
        else:
            c.fill = rdn_mid_fill

        # E: PV
        if pv_arr is not None and i < len(pv_arr):
            ws.cell(row=r, column=5, value=round(float(pv_arr[i]), 1)).fill = pv_fill

        # F: Load
        if load_arr is not None and i < len(load_arr):
            ws.cell(row=r, column=6, value=round(float(load_arr[i]), 1))

        # G: SoC [kWh]
        c = ws.cell(row=r, column=7, value=round(soc, 1))
        c.fill = batt_fill

        # H: SoC [%] — formula
        c = ws.cell(row=r, column=8, value=f'=G{r}/{bess_energy_kwh}*100')
        c.fill = formula_fill
        c.number_format = '0.0'

        # I: Charge [kW]
        c = ws.cell(row=r, column=9, value=round(ch, 1))
        if ch > 0.1:
            c.fill = charge_fill
            c.font = Font(bold=True)

        # J: Discharge [kW]
        c = ws.cell(row=r, column=10, value=round(dis, 1))
        if dis > 0.1:
            c.fill = discharge_fill
            c.font = Font(bold=True)

        # K: Stan (status label)
        if ch > 0.1:
            c = ws.cell(row=r, column=11, value="ŁADOWANIE")
            c.fill = charge_fill
            c.font = Font(bold=True, color="1565C0")
        elif dis > 0.1:
            c = ws.cell(row=r, column=11, value="ROZŁADOWANIE")
            c.fill = discharge_fill
            c.font = Font(bold=True, color="E65100")
        else:
            c = ws.cell(row=r, column=11, value="—")
            c.alignment = Alignment(horizontal='center')

        # L: Koszt ładowania = charge_kw * dt_hours * price / 1000 [PLN]
        #    (energy charged * RDN price = opportunity cost of storing vs exporting)
        c = ws.cell(row=r, column=12, value=f'=I{r}*{dt_hours}*D{r}/1000')
        c.number_format = '0.00'
        if ch > 0.1:
            c.fill = charge_fill

        # M: Wartość rozładowania = discharge_kw * dt_hours * price / 1000 [PLN]
        #    (energy discharged * RDN price = avoided import cost)
        c = ws.cell(row=r, column=13, value=f'=J{r}*{dt_hours}*D{r}/1000')
        c.number_format = '0.00'
        if dis > 0.1:
            c.fill = discharge_fill

        # N: Zysk netto godzinowy = wartość rozład. - koszt ładowania
        c = ws.cell(row=r, column=14, value=f'=M{r}-L{r}')
        c.number_format = '0.00'
        c.fill = formula_fill

        # O: Zysk skumulowany dzień — SUMPRODUCT for same date
        # Use simple running sum: if same day, add to previous; else start fresh
        if i == 0 or time_index[i].date() != time_index[i - 1].date():
            c = ws.cell(row=r, column=15, value=f'=N{r}')
        else:
            c = ws.cell(row=r, column=15, value=f'=O{r-1}+N{r}')
        c.number_format = '0.00'
        c.fill = formula_fill

        # P: Zysk skumulowany miesiąc
        if i == 0 or time_index[i].month != time_index[i - 1].month:
            c = ws.cell(row=r, column=16, value=f'=N{r}')
        else:
            c = ws.cell(row=r, column=16, value=f'=P{r-1}+N{r}')
        c.number_format = '0.00'
        c.fill = formula_fill

        # Q: Zysk skumulowany rok (running total)
        if i == 0:
            c = ws.cell(row=r, column=17, value=f'=N{r}')
        else:
            c = ws.cell(row=r, column=17, value=f'=Q{r-1}+N{r}')
        c.number_format = '0.00'
        c.fill = savings_fill
        c.font = Font(bold=True)

        # Day separator — bold bottom border at 23:00
        if hour == 23:
            day_border = Border(bottom=Side(style='medium', color='999999'))
            for col in range(1, 18):
                ws.cell(row=r, column=col).border = day_border

    last_row = n_rows + 4

    # ===== Summary row 3 =====
    ws.cell(row=3, column=1, value="PODSUMOWANIE:").font = Font(bold=True, size=11)

    # Total charge cost
    c = ws.cell(row=3, column=12, value=f'=SUM(L5:L{last_row})')
    c.font = Font(bold=True)
    c.number_format = '#,##0'
    c.fill = summary_fill

    # Total discharge value
    c = ws.cell(row=3, column=13, value=f'=SUM(M5:M{last_row})')
    c.font = Font(bold=True)
    c.number_format = '#,##0'
    c.fill = summary_fill

    # Total net profit
    c = ws.cell(row=3, column=14, value=f'=SUM(N5:N{last_row})')
    c.font = Font(bold=True, size=12, color="1B5E20")
    c.number_format = '#,##0'
    c.fill = savings_fill

    # Labels in row 3
    ws.cell(row=3, column=11, value="SUMY →").font = Font(bold=True)
    ws.cell(row=3, column=9, value=f'=SUM(I5:I{last_row})*{dt_hours}').number_format = '#,##0'
    ws.cell(row=3, column=9).font = Font(bold=True)
    ws.cell(row=3, column=10, value=f'=SUM(J5:J{last_row})*{dt_hours}').number_format = '#,##0'
    ws.cell(row=3, column=10).font = Font(bold=True)

    # ===== Daily/monthly summary block below data =====
    sr = last_row + 3  # summary start row

    ws.cell(row=sr, column=1, value="PODSUMOWANIE DZIENNE I MIESIĘCZNE").font = Font(bold=True, size=13)
    ws.merge_cells(f'A{sr}:F{sr}')
    sr += 1

    # Daily summary header
    day_headers = [
        ('Data', 12), ('Dzień tyg.', 8), ('Ładowanie [kWh]', 16),
        ('Rozładowanie [kWh]', 16), ('Koszt ładow. [PLN]', 16),
        ('Wartość rozład. [PLN]', 16), ('ZYSK [PLN]', 14),
        ('Śr. cena ładow.\n[PLN/MWh]', 16), ('Śr. cena rozład.\n[PLN/MWh]', 16),
        ('Spread\n[PLN/MWh]', 14),
    ]

    for ci, (h, w) in enumerate(day_headers, 1):
        cell = ws.cell(row=sr, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = max(
            ws.column_dimensions[get_column_letter(ci)].width or 0, w
        )
    sr += 1

    # Aggregate by day
    from collections import defaultdict
    daily = defaultdict(lambda: {
        'charge_kwh': 0, 'discharge_kwh': 0,
        'charge_cost': 0, 'discharge_value': 0,
        'charge_price_sum': 0, 'charge_hours': 0,
        'discharge_price_sum': 0, 'discharge_hours': 0,
        'weekday': 0,
    })

    for i in range(n_rows):
        d = time_index[i].date()
        ch = float(charge_arr[i]) if charge_arr is not None and i < len(charge_arr) else 0
        dis = float(discharge_arr[i]) if discharge_arr is not None and i < len(discharge_arr) else 0
        price = float(rdn[i])

        daily[d]['weekday'] = time_index[i].weekday()
        if ch > 0.1:
            daily[d]['charge_kwh'] += ch * dt_hours
            daily[d]['charge_cost'] += ch * dt_hours * price / 1000
            daily[d]['charge_price_sum'] += price
            daily[d]['charge_hours'] += 1
        if dis > 0.1:
            daily[d]['discharge_kwh'] += dis * dt_hours
            daily[d]['discharge_value'] += dis * dt_hours * price / 1000
            daily[d]['discharge_price_sum'] += price
            daily[d]['discharge_hours'] += 1

    daily_start_row = sr
    for d in sorted(daily.keys()):
        dd = daily[d]
        avg_ch_price = dd['charge_price_sum'] / dd['charge_hours'] if dd['charge_hours'] > 0 else 0
        avg_dis_price = dd['discharge_price_sum'] / dd['discharge_hours'] if dd['discharge_hours'] > 0 else 0
        profit = dd['discharge_value'] - dd['charge_cost']

        ws.cell(row=sr, column=1, value=d)
        ws.cell(row=sr, column=2, value=weekdays_pl[dd['weekday']])
        ws.cell(row=sr, column=3, value=round(dd['charge_kwh'], 1)).number_format = '0.0'
        ws.cell(row=sr, column=4, value=round(dd['discharge_kwh'], 1)).number_format = '0.0'
        ws.cell(row=sr, column=5, value=round(dd['charge_cost'], 2)).number_format = '0.00'
        ws.cell(row=sr, column=6, value=round(dd['discharge_value'], 2)).number_format = '0.00'

        c = ws.cell(row=sr, column=7, value=round(profit, 2))
        c.number_format = '0.00'
        c.font = Font(bold=True)
        c.fill = profit_fill if profit >= 0 else loss_fill

        ws.cell(row=sr, column=8, value=round(avg_ch_price, 0)).number_format = '0'
        ws.cell(row=sr, column=9, value=round(avg_dis_price, 0)).number_format = '0'

        spread = avg_dis_price - avg_ch_price
        c = ws.cell(row=sr, column=10, value=round(spread, 0))
        c.number_format = '0'
        c.font = Font(bold=True, color="1B5E20" if spread > 0 else "B71C1C")

        sr += 1

    # Daily totals
    c = ws.cell(row=sr, column=1, value="SUMA DZIENNYCH")
    c.font = Font(bold=True, size=11)
    for col in range(3, 8):
        letter = get_column_letter(col)
        c = ws.cell(row=sr, column=col, value=f'=SUM({letter}{daily_start_row}:{letter}{sr-1})')
        c.font = Font(bold=True, size=11)
        c.number_format = '#,##0'
        c.fill = summary_fill
    sr += 2

    # ===== Monthly summary =====
    ws.cell(row=sr, column=1, value="PODSUMOWANIE MIESIĘCZNE").font = Font(bold=True, size=13)
    ws.merge_cells(f'A{sr}:F{sr}')
    sr += 1

    month_headers = ['Miesiąc', 'Ładowanie [kWh]', 'Rozładowanie [kWh]',
                     'Koszt ładow. [PLN]', 'Wartość rozład. [PLN]', 'ZYSK [PLN]',
                     'Śr. cena ładow.', 'Śr. cena rozład.', 'Spread']
    for ci, h in enumerate(month_headers, 1):
        cell = ws.cell(row=sr, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    sr += 1

    monthly = defaultdict(lambda: {
        'charge_kwh': 0, 'discharge_kwh': 0,
        'charge_cost': 0, 'discharge_value': 0,
        'charge_price_sum': 0, 'charge_hours': 0,
        'discharge_price_sum': 0, 'discharge_hours': 0,
    })

    for i in range(n_rows):
        m = time_index[i].month
        ch = float(charge_arr[i]) if charge_arr is not None and i < len(charge_arr) else 0
        dis = float(discharge_arr[i]) if discharge_arr is not None and i < len(discharge_arr) else 0
        price = float(rdn[i])

        if ch > 0.1:
            monthly[m]['charge_kwh'] += ch * dt_hours
            monthly[m]['charge_cost'] += ch * dt_hours * price / 1000
            monthly[m]['charge_price_sum'] += price
            monthly[m]['charge_hours'] += 1
        if dis > 0.1:
            monthly[m]['discharge_kwh'] += dis * dt_hours
            monthly[m]['discharge_value'] += dis * dt_hours * price / 1000
            monthly[m]['discharge_price_sum'] += price
            monthly[m]['discharge_hours'] += 1

    month_names = ['', 'Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec',
                   'Lipiec', 'Sierpień', 'Wrzesień', 'Październik', 'Listopad', 'Grudzień']
    monthly_start_row = sr
    for m in sorted(monthly.keys()):
        mm = monthly[m]
        avg_ch = mm['charge_price_sum'] / mm['charge_hours'] if mm['charge_hours'] > 0 else 0
        avg_dis = mm['discharge_price_sum'] / mm['discharge_hours'] if mm['discharge_hours'] > 0 else 0
        profit = mm['discharge_value'] - mm['charge_cost']

        ws.cell(row=sr, column=1, value=month_names[m]).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=round(mm['charge_kwh'], 0)).number_format = '#,##0'
        ws.cell(row=sr, column=3, value=round(mm['discharge_kwh'], 0)).number_format = '#,##0'
        ws.cell(row=sr, column=4, value=round(mm['charge_cost'], 0)).number_format = '#,##0'
        ws.cell(row=sr, column=5, value=round(mm['discharge_value'], 0)).number_format = '#,##0'

        c = ws.cell(row=sr, column=6, value=round(profit, 0))
        c.number_format = '#,##0'
        c.font = Font(bold=True, size=11)
        c.fill = profit_fill if profit >= 0 else loss_fill

        ws.cell(row=sr, column=7, value=round(avg_ch, 0)).number_format = '0'
        ws.cell(row=sr, column=8, value=round(avg_dis, 0)).number_format = '0'
        c = ws.cell(row=sr, column=9, value=round(avg_dis - avg_ch, 0))
        c.number_format = '0'
        c.font = Font(bold=True, color="1B5E20" if (avg_dis - avg_ch) > 0 else "B71C1C")

        sr += 1

    # Monthly totals
    c = ws.cell(row=sr, column=1, value="SUMA ROCZNA")
    c.font = Font(bold=True, size=12)
    for col in range(2, 7):
        letter = get_column_letter(col)
        c = ws.cell(row=sr, column=col, value=f'=SUM({letter}{monthly_start_row}:{letter}{sr-1})')
        c.font = Font(bold=True, size=12)
        c.number_format = '#,##0'
        c.fill = savings_fill

    ws.freeze_panes = 'A5'
