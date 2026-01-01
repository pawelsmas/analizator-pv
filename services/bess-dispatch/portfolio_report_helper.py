"""
Portfolio Report Helper (v2.4.0)

Creates ZIP bundles for portfolio reports.

ZIP contains:
- portfolio_summary.json: Aggregated portfolio summary
- portfolio_items.csv: Items CSV export
- portfolio_items.xlsx: Items Excel export (optional)
- README.md: Documentation and repro instructions
- WARNINGS.txt: Warnings (if any)
"""

import json
import zipfile
import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from report_models import ReportOptions, ReportBranding


# Fixed ZIP timestamp for determinism
FIXED_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)


def build_portfolio_report_zip(
    portfolio_summary: Dict[str, Any],
    portfolio_items: List[Dict[str, Any]],
    options: Optional[ReportOptions] = None,
    source_id: Optional[str] = None,
    source_type: str = "api",
) -> bytes:
    """
    Build a deterministic ZIP bundle for portfolio report.

    Args:
        portfolio_summary: Aggregated portfolio summary dict
        portfolio_items: List of portfolio item dicts
        options: Report generation options
        source_id: Source identifier (job_id or request_id)
        source_type: "api" or "job"

    Returns:
        ZIP file as bytes
    """
    if options is None:
        options = ReportOptions()

    # Collect warnings
    warnings = []

    # Check for mixed assumptions
    if portfolio_summary.get("mixed_assumptions"):
        warnings.append("Mixed assumptions versions detected across runs")

    # Prepare files for ZIP
    files = {}

    # 1. portfolio_summary.json
    files["portfolio_summary.json"] = _serialize_json(portfolio_summary)

    # 2. portfolio_items.csv
    files["portfolio_items.csv"] = _build_items_csv(portfolio_items)

    # 3. portfolio_items.xlsx (optional - as bytes, not string)
    xlsx_bytes = _build_items_xlsx(portfolio_items)

    # 4. README.md
    files["README.md"] = _build_readme(portfolio_summary, options, source_id, source_type)

    # 5. WARNINGS.txt (if any)
    if warnings:
        files["WARNINGS.txt"] = _build_warnings_file(warnings)

    # Build deterministic ZIP
    return _build_zip(files, xlsx_bytes)


def _serialize_json(data: Any) -> str:
    """Serialize data to deterministic JSON string."""
    return json.dumps(
        data,
        sort_keys=True,
        ensure_ascii=False,
        indent=2,
    )


def _build_items_csv(items: List[Dict[str, Any]]) -> str:
    """Build CSV content from portfolio items."""
    if not items:
        return "run_id,label,status,recommended_variant,npv_pln,payback_years,irr_pct,net_savings_pln,capex_pln\n"

    # Header
    columns = [
        "run_id", "label", "status", "recommended_variant", "objective_used",
        "npv_pln", "payback_years", "irr_pct", "net_savings_pln", "capex_pln",
        "assumptions_version", "schema_version", "pricing_mode", "tags", "error_message"
    ]

    lines = [",".join(columns)]

    # Data rows
    for item in items:
        row = []
        for col in columns:
            value = item.get(col)
            if value is None:
                row.append("")
            elif col == "tags" and isinstance(value, list):
                # Join tags with semicolon (avoid CSV comma issues)
                row.append(";".join(str(v) for v in value))
            elif isinstance(value, (int, float)):
                row.append(str(value))
            else:
                # Quote strings that might contain commas
                str_val = str(value).replace('"', '""')
                if "," in str_val or '"' in str_val or "\n" in str_val:
                    row.append(f'"{str_val}"')
                else:
                    row.append(str_val)
        lines.append(",".join(row))

    return "\n".join(lines)


def _build_items_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """Build XLSX content from portfolio items."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio Items"

        # Header
        columns = [
            "run_id", "label", "status", "recommended_variant", "objective_used",
            "npv_pln", "payback_years", "irr_pct", "net_savings_pln", "capex_pln",
            "assumptions_version", "schema_version", "pricing_mode"
        ]

        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        # Data rows
        for row_idx, item in enumerate(items, 2):
            for col_idx, col in enumerate(columns, 1):
                value = item.get(col)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-width columns
        for col in range(1, len(columns) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except ImportError:
        # openpyxl not available
        return b""


def _build_readme(
    portfolio_summary: Dict[str, Any],
    options: ReportOptions,
    source_id: Optional[str],
    source_type: str,
) -> str:
    """Build README.md with documentation."""
    branding = options.branding if options else None

    title = "Portfolio Report"
    if branding and branding.report_title:
        title = branding.report_title

    lines = [
        f"# {title}",
        "",
        "## Portfolio Summary",
        "",
        f"- **Items Total**: {portfolio_summary.get('items_total', 0)}",
        f"- **Items OK**: {portfolio_summary.get('items_ok', 0)}",
        f"- **Items Error**: {portfolio_summary.get('items_error', 0)}",
        f"- **Mixed Assumptions**: {'Yes' if portfolio_summary.get('mixed_assumptions') else 'No'}",
        "",
        "### Financial Summary",
        "",
        f"- **Total NPV**: {portfolio_summary.get('total_npv_pln', 0):,.0f} PLN",
        f"- **Total Net Savings**: {portfolio_summary.get('total_net_savings_pln', 0):,.0f} PLN",
        f"- **Total CAPEX**: {portfolio_summary.get('total_capex_pln', 0):,.0f} PLN",
    ]

    weighted_payback = portfolio_summary.get('weighted_payback_years')
    if weighted_payback:
        lines.append(f"- **Weighted Payback**: {weighted_payback:.1f} years")

    avg_irr = portfolio_summary.get('avg_irr_pct')
    if avg_irr:
        lines.append(f"- **Average IRR**: {avg_irr:.1f}%")

    lines.extend([
        "",
        "## Branding",
        "",
    ])

    if branding:
        if branding.client_name:
            lines.append(f"- **Client**: {branding.client_name}")
        if branding.site_name:
            lines.append(f"- **Site**: {branding.site_name}")
        if branding.prepared_by:
            lines.append(f"- **Prepared By**: {branding.prepared_by}")
        if branding.prepared_for:
            lines.append(f"- **Prepared For**: {branding.prepared_for}")
    else:
        lines.append("No branding specified.")

    lines.extend([
        "",
        "## Files in this bundle",
        "",
        "| File | Description |",
        "|------|-------------|",
        "| `portfolio_summary.json` | Aggregated portfolio summary |",
        "| `portfolio_items.csv` | Items as CSV |",
        "| `portfolio_items.xlsx` | Items as Excel |",
        "| `README.md` | This file |",
        "",
        "## How to get individual run reports",
        "",
        "For each run_id in portfolio_items.csv:",
        "",
        "```bash",
        "# Get report ZIP",
        "curl -O http://localhost:8031/api/bess-dispatch/runs/{run_id}/report.zip",
        "",
        "# Get report PDF",
        "curl -O http://localhost:8031/api/bess-dispatch/runs/{run_id}/report.pdf",
        "",
        "# Get report XLSX",
        "curl -O http://localhost:8031/api/bess-dispatch/runs/{run_id}/report.xlsx",
        "```",
        "",
        "---",
        f"Generated at: {datetime.utcnow().isoformat()}Z",
    ])

    if source_id:
        lines.append(f"Source: {source_type} {source_id}")

    return "\n".join(lines)


def _build_warnings_file(warnings: list) -> str:
    """Build WARNINGS.txt content."""
    lines = [
        "# Portfolio Report Warnings",
        "",
    ]

    for i, warning in enumerate(warnings, 1):
        lines.append(f"{i}. {warning}")

    return "\n".join(lines)


def _build_zip(files: Dict[str, str], xlsx_bytes: bytes = b"") -> bytes:
    """Build deterministic ZIP from file dict."""
    buffer = io.BytesIO()

    # Sort files alphabetically for determinism
    sorted_files = sorted(files.items(), key=lambda x: x[0])

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, content in sorted_files:
            info = zipfile.ZipInfo(filename, date_time=FIXED_ZIP_TIMESTAMP)
            info.compress_type = zipfile.ZIP_DEFLATED
            zf.writestr(info, content.encode("utf-8"))

        # Add XLSX as binary
        if xlsx_bytes:
            info = zipfile.ZipInfo("portfolio_items.xlsx", date_time=FIXED_ZIP_TIMESTAMP)
            info.compress_type = zipfile.ZIP_DEFLATED
            zf.writestr(info, xlsx_bytes)

    return buffer.getvalue()
