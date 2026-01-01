"""
Report XLSX Renderer (v2.4.0)

Generates Excel reports from ReportData using openpyxl.

Worksheets:
1. Summary: Overview with KPIs
2. Ledger: Annual costs breakdown
3. Flows: Energy flows summary
4. Constraints: Grid constraints
5. Engineering (optional): Cycles, invariants, debug events
6. Warnings: Any extraction warnings
"""

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from report_data import ReportData
from report_models import ReportProfile


# Style constants
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def render_xlsx(report_data: ReportData) -> bytes:
    """
    Render ReportData to XLSX bytes.

    Args:
        report_data: ReportData object with all sections

    Returns:
        XLSX file as bytes
    """
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Summary sheet
    _add_summary_sheet(wb, report_data)

    # 2. Ledger sheet
    _add_ledger_sheet(wb, report_data)

    # 3. Flows sheet
    _add_flows_sheet(wb, report_data)

    # 4. Constraints sheet
    _add_constraints_sheet(wb, report_data)

    # 5. Engineering sheet (if profile=engineering)
    options = report_data.options
    if options and options.profile == ReportProfile.ENGINEERING:
        _add_engineering_sheet(wb, report_data)

    # 6. Warnings sheet (if any)
    if report_data.warnings:
        _add_warnings_sheet(wb, report_data)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _style_header_row(ws, row: int, cols: int):
    """Apply header style to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, cols: int, total_row: int = None):
    """Apply data style to rows."""
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER

            if row == total_row:
                cell.fill = TOTAL_FILL
                cell.font = Font(bold=True)
            elif row % 2 == 0:
                cell.fill = ALT_ROW_FILL


def _auto_column_width(ws, cols: int):
    """Auto-adjust column widths."""
    for col in range(1, cols + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _add_summary_sheet(wb: Workbook, report_data: ReportData):
    """Add Summary worksheet."""
    ws = wb.create_sheet("Summary")

    meta = report_data.meta
    rec = report_data.recommended
    options = report_data.options
    branding = options.branding if options else None

    # Title
    title = "BESS Sizing Report"
    if branding and branding.report_title:
        title = branding.report_title

    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(size=16, bold=True)

    row = 3

    # Branding info
    if branding:
        if branding.client_name:
            ws.cell(row=row, column=1, value="Client")
            ws.cell(row=row, column=2, value=branding.client_name)
            row += 1
        if branding.site_name:
            ws.cell(row=row, column=1, value="Site")
            ws.cell(row=row, column=2, value=branding.site_name)
            row += 1
        if branding.prepared_by:
            ws.cell(row=row, column=1, value="Prepared By")
            ws.cell(row=row, column=2, value=branding.prepared_by)
            row += 1
        if branding.prepared_for:
            ws.cell(row=row, column=1, value="Prepared For")
            ws.cell(row=row, column=2, value=branding.prepared_for)
            row += 1

    row += 1

    # Run info
    ws.cell(row=row, column=1, value="Run ID")
    ws.cell(row=row, column=2, value=meta.run_id)
    row += 1

    ws.cell(row=row, column=1, value="Created")
    ws.cell(row=row, column=2, value=meta.created_at)
    row += 1

    if meta.schema_version:
        ws.cell(row=row, column=1, value="Schema Version")
        ws.cell(row=row, column=2, value=meta.schema_version)
        row += 1

    if meta.assumptions_version:
        ws.cell(row=row, column=1, value="Assumptions")
        ws.cell(row=row, column=2, value=meta.assumptions_version)
        row += 1

    row += 1

    # Period info
    if meta.period_start:
        ws.cell(row=row, column=1, value="Period Start")
        ws.cell(row=row, column=2, value=meta.period_start)
        row += 1

    if meta.period_end:
        ws.cell(row=row, column=1, value="Period End")
        ws.cell(row=row, column=2, value=meta.period_end)
        row += 1

    if meta.period_hours:
        ws.cell(row=row, column=1, value="Period Hours")
        ws.cell(row=row, column=2, value=meta.period_hours)
        row += 1

        ws.cell(row=row, column=1, value="Full Year")
        ws.cell(row=row, column=2, value="Yes" if meta.is_full_year else "No")
        row += 1

    row += 2

    # KPIs header
    ws.cell(row=row, column=1, value="Key Performance Indicators")
    ws.cell(row=row, column=1).font = Font(size=14, bold=True)
    row += 2

    # KPIs table
    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    _style_header_row(ws, row, 2)
    row += 1

    kpis = [
        ("Recommended Variant", rec.recommended_variant or "N/A"),
        ("Objective Used", rec.objective_used or "N/A"),
        ("NPV (PLN)", rec.npv_pln if rec.npv_pln else "N/A"),
        ("Payback (years)", rec.payback_years if rec.payback_years else "N/A"),
        ("IRR (%)", rec.irr_pct if rec.irr_pct else "N/A"),
        ("Net Savings (PLN)", rec.net_savings_pln if rec.net_savings_pln else "N/A"),
    ]

    start_row = row
    for metric, value in kpis:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=value)
        row += 1

    _style_data_rows(ws, start_row, row - 1, 2)
    _auto_column_width(ws, 2)


def _add_ledger_sheet(wb: Workbook, report_data: ReportData):
    """Add Ledger worksheet."""
    ws = wb.create_sheet("Ledger")

    ledger = report_data.money_ledger_annual

    # Header
    ws.cell(row=1, column=1, value="Annual Cost Breakdown")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    # Table header
    headers = ["Category", "Baseline (PLN)", "Project (PLN)", "Savings (PLN)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    _style_header_row(ws, 3, 4)

    # Data rows
    categories = [
        ("Import Energy", ledger.baseline.import_energy_cost_pln, ledger.project.import_energy_cost_pln, ledger.delta.import_energy_cost_pln),
        ("Export Revenue", ledger.baseline.export_revenue_pln, ledger.project.export_revenue_pln, ledger.delta.export_revenue_pln),
        ("Other Fees", ledger.baseline.other_fees_pln, ledger.project.other_fees_pln, ledger.delta.other_fees_pln),
        ("Capacity Fee", ledger.baseline.capacity_fee_pln, ledger.project.capacity_fee_pln, ledger.delta.capacity_fee_pln),
        ("Demand Charge", ledger.baseline.demand_charge_pln, ledger.project.demand_charge_pln, ledger.delta.demand_charge_pln),
        ("Unserved Penalty", ledger.baseline.unserved_penalty_pln, ledger.project.unserved_penalty_pln, ledger.delta.unserved_penalty_pln),
        ("Degradation", ledger.baseline.degradation_cost_pln, ledger.project.degradation_cost_pln, ledger.delta.degradation_cost_pln),
        ("Total Cost", ledger.baseline.total_cost_pln, ledger.project.total_cost_pln, ledger.delta.total_cost_pln),
    ]

    row = 4
    start_row = row
    for cat, base, proj, delta in categories:
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=base)
        ws.cell(row=row, column=3, value=proj)
        ws.cell(row=row, column=4, value=delta)
        row += 1

    _style_data_rows(ws, start_row, row - 1, 4, total_row=row - 1)
    _auto_column_width(ws, 4)


def _add_flows_sheet(wb: Workbook, report_data: ReportData):
    """Add Flows worksheet."""
    ws = wb.create_sheet("Flows")

    flows = report_data.flows_annual

    # Header
    ws.cell(row=1, column=1, value="Energy Flows (Annual)")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    # Table header
    ws.cell(row=3, column=1, value="Flow")
    ws.cell(row=3, column=2, value="Value")
    ws.cell(row=3, column=3, value="Unit")
    _style_header_row(ws, 3, 3)

    # Data rows
    flow_data = [
        ("Grid Import", flows.grid_import_mwh, "MWh"),
        ("Grid Export", flows.grid_export_mwh, "MWh"),
        ("PV Curtailment", flows.pv_curtail_mwh, "MWh"),
        ("Unserved Load", flows.unserved_load_kwh, "kWh"),
        ("Unserved Penalty", flows.unserved_penalty_pln, "PLN"),
    ]

    row = 4
    start_row = row
    for name, value, unit in flow_data:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=unit)
        row += 1

    _style_data_rows(ws, start_row, row - 1, 3)
    _auto_column_width(ws, 3)


def _add_constraints_sheet(wb: Workbook, report_data: ReportData):
    """Add Constraints worksheet."""
    ws = wb.create_sheet("Constraints")

    constraints = report_data.constraints

    # Header
    ws.cell(row=1, column=1, value="Grid Constraints")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    # Table header
    ws.cell(row=3, column=1, value="Constraint")
    ws.cell(row=3, column=2, value="Value")
    ws.cell(row=3, column=3, value="Unit")
    _style_header_row(ws, 3, 3)

    # Data rows
    constraint_data = [
        ("Max Import", constraints.max_import_kw, "kW"),
        ("Max Export", constraints.max_export_kw, "kW"),
        ("Export Limited Steps", constraints.export_limited_steps, "steps"),
        ("Export Limited Energy", constraints.export_limited_mwh, "MWh"),
        ("Import Limited Steps", constraints.import_limited_steps, "steps"),
        ("Import Limited Energy", constraints.import_limited_mwh, "MWh"),
    ]

    row = 4
    start_row = row
    for name, value, unit in constraint_data:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=value if value is not None else "N/A")
        ws.cell(row=row, column=3, value=unit)
        row += 1

    _style_data_rows(ws, start_row, row - 1, 3)
    _auto_column_width(ws, 3)


def _add_engineering_sheet(wb: Workbook, report_data: ReportData):
    """Add Engineering worksheet."""
    ws = wb.create_sheet("Engineering")

    dispatch = report_data.dispatch_summary

    # Header
    ws.cell(row=1, column=1, value="Engineering Details")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    row = 3

    # Cycle Summary
    if dispatch and dispatch.cycle_summary:
        ws.cell(row=row, column=1, value="Cycle Summary")
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        row += 1

        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value="Value")
        _style_header_row(ws, row, 2)
        row += 1

        cycle = dispatch.cycle_summary
        cycle_data = [
            ("Total EFC", cycle.efc_total),
            ("Avg Daily EFC", cycle.cycles_per_day_avg),
            ("Max Daily EFC", cycle.cycles_per_day_max),
            ("Total Throughput (kWh)", cycle.total_throughput_kwh),
            ("Cycle Limit Hit", "Yes" if cycle.cycle_limit_hit else "No"),
        ]

        start_row = row
        for name, value in cycle_data:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=value if value is not None else "N/A")
            row += 1

        _style_data_rows(ws, start_row, row - 1, 2)
        row += 2

    # Invariants
    if dispatch and dispatch.invariants:
        ws.cell(row=row, column=1, value="Invariant Checks")
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        row += 1

        ws.cell(row=row, column=1, value="Check")
        ws.cell(row=row, column=2, value="Status")
        _style_header_row(ws, row, 2)
        row += 1

        inv = dispatch.invariants
        status = lambda ok: "PASS" if ok else "FAIL"
        inv_data = [
            ("All OK", status(inv.all_ok)),
            ("SOC Bounds", status(inv.soc_bounds_ok)),
            ("Power Limits", status(inv.power_limits_ok)),
            ("No Simultaneous", status(inv.no_simultaneous_ok)),
            ("Energy Balance", status(inv.energy_balance_ok)),
        ]

        start_row = row
        for name, value in inv_data:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=value)
            row += 1

        _style_data_rows(ws, start_row, row - 1, 2)
        row += 2

    # Debug Events
    if dispatch and dispatch.debug_events:
        ws.cell(row=row, column=1, value="Debug Events")
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        row += 1

        ws.cell(row=row, column=1, value="Event")
        ws.cell(row=row, column=2, value="Value")
        _style_header_row(ws, row, 2)
        row += 1

        debug = dispatch.debug_events
        debug_data = [
            ("Total Steps", debug.steps),
            ("Export Limited Steps", debug.export_limited_steps),
            ("Export Limited (MWh)", debug.export_limited_mwh),
            ("Import Limited Steps", debug.import_limited_steps),
            ("Import Limited (MWh)", debug.import_limited_mwh),
            ("PV Curtailment (MWh)", debug.pv_curtail_mwh),
            ("Unserved Load (kWh)", debug.unserved_load_kwh),
        ]

        start_row = row
        for name, value in debug_data:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=value)
            row += 1

        _style_data_rows(ws, start_row, row - 1, 2)

    _auto_column_width(ws, 2)


def _add_warnings_sheet(wb: Workbook, report_data: ReportData):
    """Add Warnings worksheet."""
    ws = wb.create_sheet("Warnings")

    # Header
    ws.cell(row=1, column=1, value="Report Extraction Warnings")
    ws.cell(row=1, column=1).font = Font(size=14, bold=True)

    ws.cell(row=3, column=1, value="#")
    ws.cell(row=3, column=2, value="Warning")
    _style_header_row(ws, 3, 2)

    row = 4
    start_row = row
    for i, warning in enumerate(report_data.warnings, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=warning)
        row += 1

    _style_data_rows(ws, start_row, row - 1, 2)
    _auto_column_width(ws, 2)
